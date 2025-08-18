# intern.py
# Lógica completa para el área de planificación internacional, incluyendo modelos y exportaciones.

import os
import json
# Se elimina la importación de Decimal ya que no se usará
from datetime import datetime
from io import BytesIO
from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, jsonify, Response, send_file, abort)
from werkzeug.utils import secure_filename

# --- LIBRERÍAS PARA EXPORTACIÓN ---
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

# --- CONFIGURACIÓN DEL BLUEPRINT ---
intern_bp = Blueprint('intern', __name__,
                        template_folder='templates',
                        static_folder='static')

UPLOAD_FOLDER = 'static/uploads/intern_flyers'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# --- SIMULACIÓN DE BASE DE DATOS ---
VIAJES_DB = [
    {
        "id": "1",
        "nombre_viaje": "Aventura en Bocas del Toro",
        "pais": "Panamá",
        "flyer_url": "uploads/intern_flyers/panama_flyer_ejemplo.jpg",
        "fecha_viaje": "2024-12-15",
        "fecha_regreso": "2024-12-22",
        "cantidad_dias": 8,
        # CORRECCIÓN: Se usan números enteros o flotantes en lugar de Decimal
        "precio_viaje_usd": 1250,
        "capacidad_sg": 20,
        "capacidad_cg": 25,
        "tipo_moneda_base": "USD",
        "tipo_cambio_a_crc": 520,
        "pagos_y_servicios": [
            {
                "tipo_pago": "Estadía",
                "nombre_contacto": "Hotel Mar y Sol",
                "cantidad_dias_evento": "5",
                "precio": "100",
                "poliza": "Si",
                "telefonos": [
                    {"tipo": "WHATSAPP", "codigo": "+507", "numero": "1234-5678", "contacto": "Recepción"}
                ]
            }
        ],
        "calculos": {
            "total_neto_usd": 500.0,
            "precio_individual_usd": 25.0
        }
    }
]
next_id = 2

# --- FUNCIONES AUXILIARES ---
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def find_viaje_by_id(viaje_id):
    for viaje in VIAJES_DB:
        if viaje['id'] == viaje_id:
            return viaje
    return None

# --- RUTAS PRINCIPALES (CRUD) ---

@intern_bp.route('/crear', methods=['GET', 'POST'])
def crear_intern():
    global next_id
    if request.method == 'POST':
        flyer_filename = None
        flyer_url_for_db = None

        if 'flyer' in request.files:
            file = request.files['flyer']
            if file and file.filename != '' and allowed_file(file.filename):
                flyer_filename = secure_filename(file.filename)
                os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                file.save(os.path.join(UPLOAD_FOLDER, flyer_filename))
                flyer_url_for_db = f"uploads/intern_flyers/{flyer_filename}"

        pagos_data = {}
        for key, value in request.form.items():
            if key.startswith('pagos['):
                parts = key.replace(']', '').split('[')
                pago_index = parts[1]
                if pago_index not in pagos_data:
                    pagos_data[pago_index] = {}
                if len(parts) > 3:
                    nested_type, nested_id, nested_field = parts[2], parts[3], parts[4]
                    if nested_type not in pagos_data[pago_index]:
                        pagos_data[pago_index][nested_type] = {}
                    if nested_id not in pagos_data[pago_index][nested_type]:
                        pagos_data[pago_index][nested_type][nested_id] = {}
                    pagos_data[pago_index][nested_type][nested_id][nested_field] = value
                else:
                    field_name = parts[2]
                    pagos_data[pago_index][field_name] = value
        
        pagos_procesados = []
        for index, pago_info in pagos_data.items():
            for nested_type in ['telefonos', 'emails', 'urls', 'servicios', 'fechas']:
                if nested_type in pago_info and isinstance(pago_info[nested_type], dict):
                    pago_info[nested_type] = list(pago_info[nested_type].values())
            pagos_procesados.append(pago_info)
        
        # CORRECCIÓN: Se usan flotantes para los cálculos
        total_neto_usd = 0.0
        for pago in pagos_procesados:
            cantidad = float(pago.get('cantidad_dias_evento', '0'))
            precio = float(pago.get('precio', '0'))
            total_neto_usd += cantidad * precio
        
        capacidad_sg = float(request.form.get('capacidad_sg', '1'))
        precio_individual_usd = total_neto_usd / capacidad_sg if capacidad_sg > 0 else 0.0

        nuevo_viaje = {
            "id": str(next_id),
            "nombre_viaje": request.form.get('nombre_viaje'),
            "pais": request.form.get('pais'),
            "flyer_url": flyer_url_for_db,
            "fecha_viaje": request.form.get('fecha_viaje'),
            "fecha_regreso": request.form.get('fecha_regreso'),
            "cantidad_dias": request.form.get('cantidad_dias', type=int),
            # CORRECCIÓN: Se convierte a float
            "precio_viaje_usd": request.form.get('precio_viaje', type=float),
            "capacidad_sg": request.form.get('capacidad_sg', type=int),
            "capacidad_cg": request.form.get('capacidad_cg', type=int),
            "tipo_moneda_base": request.form.get('tipo_moneda'),
            "tipo_cambio_a_crc": request.form.get('tipo_cambio', type=float),
            "pagos_y_servicios": pagos_procesados,
            "calculos": {
                "total_neto_usd": total_neto_usd,
                "precio_individual_usd": precio_individual_usd
            }
        }
        
        VIAJES_DB.append(nuevo_viaje)
        next_id += 1
        flash(f'El viaje "{nuevo_viaje["nombre_viaje"]}" ha sido creado.', 'success')
        return redirect(url_for('intern.ver_intern'))
        
    return render_template('crear_intern.html')

@intern_bp.route('/')
def ver_intern():
    return render_template('ver_intern.html', viajes=VIAJES_DB)

@intern_bp.route('/<viaje_id>')
def detalle_intern(viaje_id):
    viaje = find_viaje_by_id(viaje_id)
    if not viaje: abort(404)
    return render_template('detalle_intern.html', viaje=viaje)

@intern_bp.route('/<viaje_id>/editar', methods=['GET', 'POST'])
def editar_intern(viaje_id):
    viaje = find_viaje_by_id(viaje_id)
    if not viaje:
        abort(404)

    if request.method == 'POST':
        viaje['nombre_viaje'] = request.form.get('nombre_viaje')
        viaje['pais'] = request.form.get('pais')
        viaje['fecha_viaje'] = request.form.get('fecha_viaje')
        viaje['fecha_regreso'] = request.form.get('fecha_regreso')
        viaje['cantidad_dias'] = request.form.get('cantidad_dias', type=int)
        # CORRECCIÓN: Se convierte a float
        viaje['precio_viaje_usd'] = request.form.get('precio_viaje', type=float)
        viaje['capacidad_sg'] = request.form.get('capacidad_sg', type=int)
        viaje['capacidad_cg'] = request.form.get('capacidad_cg', type=int)
        viaje['tipo_moneda_base'] = request.form.get('tipo_moneda')
        viaje['tipo_cambio_a_crc'] = request.form.get('tipo_cambio', type=float)

        if 'flyer' in request.files:
            file = request.files['flyer']
            if file and file.filename != '' and allowed_file(file.filename):
                flyer_filename = secure_filename(file.filename)
                os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                file.save(os.path.join(UPLOAD_FOLDER, flyer_filename))
                viaje['flyer_url'] = f"uploads/intern_flyers/{flyer_filename}"

        pagos_data = {}
        for key, value in request.form.items():
            if key.startswith('pagos['):
                parts = key.replace(']', '').split('[')
                pago_index = parts[1]
                if pago_index not in pagos_data:
                    pagos_data[pago_index] = {}
                if len(parts) > 3:
                    nested_type, nested_id, nested_field = parts[2], parts[3], parts[4]
                    if nested_type not in pagos_data[pago_index]:
                        pagos_data[pago_index][nested_type] = {}
                    if nested_id not in pagos_data[pago_index][nested_type]:
                        pagos_data[pago_index][nested_type][nested_id] = {}
                    pagos_data[pago_index][nested_type][nested_id][nested_field] = value
                else:
                    field_name = parts[2]
                    pagos_data[pago_index][field_name] = value
        
        pagos_procesados = []
        for index, pago_info in pagos_data.items():
            for nested_type in ['telefonos', 'emails', 'urls', 'servicios', 'fechas']:
                if nested_type in pago_info and isinstance(pago_info[nested_type], dict):
                    pago_info[nested_type] = list(pago_info[nested_type].values())
            pagos_procesados.append(pago_info)
        
        viaje['pagos_y_servicios'] = pagos_procesados

        # CORRECCIÓN: Se usan flotantes para los cálculos
        total_neto_usd = 0.0
        for pago in pagos_procesados:
            cantidad = float(pago.get('cantidad_dias_evento', '0'))
            precio = float(pago.get('precio', '0'))
            total_neto_usd += cantidad * precio
        
        capacidad_sg = float(viaje.get('capacidad_sg', '1'))
        precio_individual_usd = total_neto_usd / capacidad_sg if capacidad_sg > 0 else 0.0
        
        viaje['calculos'] = {
            "total_neto_usd": total_neto_usd,
            "precio_individual_usd": precio_individual_usd
        }
        
        flash(f'El viaje "{viaje["nombre_viaje"]}" ha sido actualizado.', 'success')
        return redirect(url_for('intern.detalle_intern', viaje_id=viaje_id))

    return render_template('editar_intern.html', viaje=viaje)


@intern_bp.route('/<viaje_id>/borrar', methods=['POST'])
def borrar_intern(viaje_id):
    global VIAJES_DB
    viaje = find_viaje_by_id(viaje_id)
    if not viaje: abort(404)
    VIAJES_DB = [v for v in VIAJES_DB if v['id'] != viaje_id]
    flash(f'El viaje "{viaje["nombre_viaje"]}" ha sido eliminado.', 'warning')
    return redirect(url_for('intern.ver_intern'))

# --- RUTAS DE EXPORTACIÓN ---
@intern_bp.route('/<viaje_id>/exportar/txt')
def exportar_txt(viaje_id):
    viaje = find_viaje_by_id(viaje_id)
    if not viaje: abort(404)
    content = f"Resumen del Viaje: {viaje.get('nombre_viaje')}"
    return Response(content, mimetype="text/plain", headers={"Content-Disposition": f"attachment;filename=viaje_{viaje_id}.txt"})

@intern_bp.route('/<viaje_id>/exportar/pdf')
def exportar_pdf(viaje_id):
    viaje = find_viaje_by_id(viaje_id)
    if not viaje: abort(404)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(40, 10, f"Viaje: {viaje.get('nombre_viaje')}")
    pdf_bytes = pdf.output(dest='S').encode('latin-1')
    return Response(pdf_bytes, mimetype='application/pdf', headers={'Content-Disposition': f'attachment;filename=viaje_{viaje_id}.pdf'})

@intern_bp.route('/<viaje_id>/exportar/excel')
def exportar_excel(viaje_id):
    viaje = find_viaje_by_id(viaje_id)
    if not viaje: abort(404)
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Nombre del Viaje"
    ws['B1'] = viaje.get('nombre_viaje')
    mem_stream = BytesIO()
    wb.save(mem_stream)
    mem_stream.seek(0)
    return send_file(mem_stream, as_attachment=True, download_name=f'viaje_{viaje_id}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@intern_bp.route('/<viaje_id>/exportar/jpg')
def exportar_jpg(viaje_id):
    viaje = find_viaje_by_id(viaje_id)
    if not viaje: abort(404)
    flyer_path = viaje.get('flyer_url')
    physical_path = os.path.join('static', flyer_path) if flyer_path else None
    if not physical_path or not os.path.exists(physical_path):
        flash('Este viaje no tiene un flyer para exportar.', 'warning')
        return redirect(url_for('intern.detalle_intern', viaje_id=viaje_id))
    return send_file(physical_path, as_attachment=True)
