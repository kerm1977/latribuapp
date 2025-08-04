from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, send_file, Response, jsonify
from models import db, Instruction, Caminata, User # Importa los modelos Instruction, Caminata y User
import os
from werkzeug.utils import secure_filename
from datetime import datetime, date, time
import json
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter # Importar para obtener la letra de la columna

# Importaciones para ReportLab (PDF)
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch # Para usar pulgadas como unidad

# Importaciones para Pillow (JPG)
from PIL import Image, ImageDraw, ImageFont

# Para limpiar HTML de campos CKEditor
from bs4 import BeautifulSoup

# Crear el Blueprint para el módulo de instrucciones
instrucciones_bp = Blueprint('instrucciones', __name__, template_folder='templates')

# NOTA: La configuración de las carpetas de subida de imágenes
# se gestiona de forma centralizada en app.py.

# Contenido predeterminado para los campos CKEditor
DEFAULT_OTRAS_RECOMENDACIONES = """
-Llevar Medicamentos personales (ES SU RESPONSABILIDAD)
-Avisar a un familiar, compañer@ o conocido que se va a rezagar momentaneamente por la razón que sea. No es nuestra obligación saber donde estarás metid@
-Desayunaremos al llegar para evitar llevar peso y así dejar las cosas en la buseta
-Siempre es opcional: pasar a comer a la salida y se decide en el camino Según La Mayoría o se programa previamente
-Evite arrecostarse a arboles, poner la mochila o prendas en el suelo y en caso de tener que hacerlo revise bien el equipo que no hayan insectos que pueden quedarse pegados y correr el riesgo de ser picados
-Si hay culebras🐍🐍 así que estar atentos.
"""

DEFAULT_NORMAS_GENERALES = """
-Ser Puntuales por respeto a l@s compañer@s que nos esperan en el camino
-Nadie se queda atrás, ni se queda solo
-NO discutir por campos de la buseta... (Ya hemos tenido problemas debido a ello y es demasiado descortés y ensucia la armonía y el ambiente bonito). Respetar el campo de su compañer@ a la venida. Conversar si se necesita hacer un cambio por la razón que sea y queda a criterio de quien puede o desea ceder ese espacio_
-NO participar en estas caminatas si se siente mal
-NO participar si tiene otras actividades en la tarde
-NO criticar al rezagado ni al más rápido
-NO alterar ni dañar el entorno (arrancar plantas, ensuciar ni extraer animalitos)
"""

DEFAULT_OTRAS_INDICACIONES_GENERALES = """
-En Caso de rayería y quedemos a la intemperie, no abracemos árboles y apaguemos todos los dispositivos que emitan señales (celulares, walkie-talkies).
-Alejese de los dispositivos eléctricos y mantenga la distancia de + de 3 metros unos a otros. No se resguarde debajo de arboles. 
-Tenga cuidado si siente algo similar a estática, es probable que un rayo caiga cerca. No se quede estático en lugares abiertos.

-Llevar Ropa de cambio y cosas personales no indispensables para caminar se quedan en la buseta
-Es probable topemos con lluvias para estar preparado con poncho aún el clima es inestable.
-EN CASO DE DERRUMBE EN EL CAMINO, INUNDACIONES O SITUACIÓN QUE PONGA EN RIESGO LA INTEGRIDAD DE TOD@S SE SUSPENDE LA CAMINATA . NO DEPENDE DE QUIENES COORDINAMOS LAS CAMINATAS. NO ES RESPONSABILIDAD DE NADIE SI DEBEMOS TOMAR OTRA RUTA MÁS LEJANA O SI EL CAMINO ESTÁ BLOQUEADO POR ACCIDENTES, LLUVIAS, O LA RAZÓN QUE SEA.)
-Si es alérgico a picaduras y/o alimentos Puede llevar una pastilla de cetirizina, fexofenadina Allegra o loratadina (OJO A LAS CONTRA-INDICACIONES MEDICAS) POR SI ES ALÉRGICO ALGUNO DE ESTOS FÁRMACOS ES MEJOR CONSULTAR ANTES DE INGERIR CUALQUIERA DE ESTOS MEDICAMENTOS.
-CADA UNO CAMINA LIBREMENTE CON NOSOTROS.*
"""


# Ruta para ver todas las instrucciones
@instrucciones_bp.route('/')
def ver_instrucciones():
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))
    
    instrucciones = Instruction.query.order_by(Instruction.fecha_creacion.desc()).all()
    return render_template('ver_instrucciones.html', instrucciones=instrucciones)

# Ruta para obtener detalles de una caminata por ID (para AJAX)
@instrucciones_bp.route('/get_caminata_details/<int:caminata_id>')
def get_caminata_details(caminata_id):
    caminata = Caminata.query.get(caminata_id)
    if caminata:
        print(f"DEBUG: Caminata encontrada: {caminata.nombre}")
        print(f"DEBUG: Caminata.fecha: {caminata.fecha}, type: {type(caminata.fecha)}")
        print(f"DEBUG: Caminata.hora_salida: {caminata.hora_salida}, type: {type(caminata.hora_salida)}")

        fecha_salida_str = ''
        # Permitir tanto date como datetime objects
        if isinstance(caminata.fecha, (date, datetime)):
            fecha_salida_str = caminata.fecha.strftime('%Y-%m-%d')

        hora_salida_str = ''
        # Permitir tanto time como datetime objects
        if isinstance(caminata.hora_salida, (time, datetime)):
            hora_salida_str = caminata.hora_salida.strftime('%H:%M')

        # Estos campos son redundantes si son iguales a fecha_salida y hora_salida, pero se mantienen por consistencia
        fecha_caminata_str = ''
        if isinstance(caminata.fecha, (date, datetime)):
            fecha_caminata_str = caminata.fecha.strftime('%Y-%m-%d')

        hora_inicio_caminata_str = ''
        if isinstance(caminata.hora_salida, (time, datetime)):
            hora_inicio_caminata_str = caminata.hora_salida.strftime('%H:%M')

        response_data = {
            'dificultad': getattr(caminata, 'dificultad', 'N/A'), # Usar getattr para manejar la ausencia del atributo
            'distancia': caminata.distancia,
            'capacidad': getattr(caminata, 'capacidad', 'N/A'), # Usar getattr para manejar la ausencia del atributo
            'lugar_salida': getattr(caminata, 'lugar_salida', 'N/A'), # Usar getattr para manejar la ausencia del atributo
            'fecha_salida': fecha_salida_str, # Usar la fecha de la caminata como fecha de salida
            'hora_salida': hora_salida_str, # Usar la hora de salida de la caminata
            'fecha_caminata': fecha_caminata_str, # Usar la fecha de la caminata como fecha de caminata
            'hora_inicio_caminata': hora_inicio_caminata_str # Usar la hora de salida de la caminata como hora de inicio
        }
        print(f"DEBUG: Datos de caminata enviados desde el backend para ID {caminata_id}: {response_data}") # Log de depuración
        return jsonify(response_data)
    print(f"DEBUG: Caminata con ID {caminata_id} no encontrada.") # Log de depuración
    return jsonify({}), 404


# Ruta para crear una nueva instrucción
@instrucciones_bp.route('/crear', methods=['GET', 'POST'])
def crear_instrucciones():
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))

    # Opciones para SELECTs
    caminatas_activas = Caminata.query.filter(Caminata.fecha >= date.today()).order_by(Caminata.fecha).all()
    dificultad_opciones = ["Iniciante", "Básico", "Intermedio","Dificil", "Avanzado", "Técnico"]
    capacidad_opciones = ["14", "17", "28", "31", "42"]
    lugar_salida_opciones = ["Parque de Tres Ríos - Escuela", "Parque de Tres Ríos - Cruz Roja", "Plaza De San Diego", "Iglesia de San Diego"]
    
    opciones_sino_opcional = ["SI", "NO", "OPCIONAL"]
    
    # Valores predeterminados para el formulario (PARA AMBOS GET Y POST)
    form_data = {
        'caminata_id': '', 
        'dificultad': '', 
        'distancia': '', 
        'capacidad': '', 
        'lugar_salida': '',
        'fecha_salida': '', 
        'hora_salida': '', 
        'fecha_caminata': '', 
        'hora_inicio_caminata': '',
        'recogemos_en': '[]', # JSON vacío para la lista de recogemos_en
        'hidratacion': '', 'litros_hidratacion': '', 'tennis_ligera': 'SI',
        'tennis_runner': '', 'tennis_hiking_baja': 'OPCIONAL', 'zapato_cana_media': 'OPCIONAL',
        'zapato_cana_alta': '', 'bastones': 'NO NECESARIOS', 'foco_headlamp': 'Siempre',
        'snacks': '', 'repelente': '', 'poncho': '', 'guantes': '', 'bloqueador': '',
        'ropa_cambio': '',
        'otras_recomendaciones': DEFAULT_OTRAS_RECOMENDACIONES,
        'normas_generales': DEFAULT_NORMAS_GENERALES,
        'otras_indicaciones_generales': DEFAULT_OTRAS_INDICACIONES_GENERALES
    }


    if request.method == 'POST':
        # Actualizar form_data con los valores del formulario
        # en caso de un error de validación para que no se pierdan los datos ingresados.
        form_data.update(request.form)
        
        caminata_id = request.form.get('caminata_id')
        dificultad = request.form.get('dificultad')
        distancia = request.form.get('distancia')
        capacidad = request.form.get('capacidad')
        lugar_salida = request.form.get('lugar_salida')
        fecha_salida_str = request.form.get('fecha_salida')
        hora_salida_str = request.form.get('hora_salida')
        fecha_caminata_str = request.form.get('fecha_caminata')
        hora_inicio_caminata_str = request.form.get('hora_inicio_caminata')
        
        recogemos_en_raw = request.form.getlist('recogemos_en[]') # Obtener como lista
        form_data['recogemos_en'] = json.dumps([item for item in recogemos_en_raw if item.strip()]) # Limpiar y a JSON


        hidratacion = request.form.get('hidratacion')
        litros_hidratacion = request.form.get('litros_hidratacion')
        tennis_ligera = request.form.get('tennis_ligera')
        tennis_runner = request.form.get('tennis_runner')
        tennis_hiking_baja = request.form.get('tennis_hiking_baja')
        zapato_cana_media = request.form.get('zapato_cana_media')
        zapato_cana_alta = request.form.get('zapato_cana_alta')
        bastones = request.form.get('bastones')
        foco_headlamp = request.form.get('foco_headlamp')
        snacks = request.form.get('snacks')
        repelente = request.form.get('repelente')
        poncho = request.form.get('poncho')
        guantes = request.form.get('guantes')
        bloqueador = request.form.get('bloqueador')
        ropa_cambio = request.form.get('ropa_cambio')

        otras_recomendaciones = request.form.get('otras_recomendaciones')
        normas_generales = request.form.get('normas_generales')
        otras_indicaciones_generales = request.form.get('otras_indicaciones_generales')

        # Validaciones de campos obligatorios
        if not all([caminata_id, dificultad, distancia, capacidad, lugar_salida,
                      fecha_salida_str, hora_salida_str]): # Removed fecha_caminata_str, hora_inicio_caminata_str
            flash('Por favor, completa todos los campos obligatorios para la información de la caminata.', 'danger')
            return render_template('crear_instrucciones.html', caminatas_activas=caminatas_activas,
                                   dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                                   lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                                   form_data=form_data) 
        
        # Conversión de tipos de datos y manejo de errores
        try:
            caminata_id = int(caminata_id)
            distancia = float(distancia) if distancia else 0.0
            fecha_salida = datetime.strptime(fecha_salida_str, '%Y-%m-%d').date()
            hora_salida = datetime.strptime(hora_salida_str, '%H:%M').time()
            # Handle optional fecha_caminata and hora_inicio_caminata
            fecha_caminata = datetime.strptime(fecha_caminata_str, '%Y-%m-%d').date() if fecha_caminata_str else None
            hora_inicio_caminata = datetime.strptime(hora_inicio_caminata_str, '%H:%M').time() if hora_inicio_caminata_str else None
        except ValueError as e:
            flash(f'Error en el formato de datos numéricos o de fecha/hora: {e}', 'danger')
            return render_template('crear_instrucciones.html', caminatas_activas=caminatas_activas,
                                   dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                                   lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                                   form_data=form_data) 

        # Validación de duplicados para fecha y hora de la caminata (similar a Calendario)
        # Only validate if both fields are provided
        if fecha_caminata and hora_inicio_caminata:
            existing_instruction_query = Instruction.query.filter_by(
                caminata_id=caminata_id,
                fecha_caminata=fecha_caminata,
                hora_inicio_caminata=hora_inicio_caminata
            )
            if existing_instruction_query.first():
                flash('Ya existe una instrucción para esta caminata en la fecha y hora de inicio especificadas. Por favor, elige otra hora o fecha para la caminata.', 'danger')
                return render_template('crear_instrucciones.html', caminatas_activas=caminatas_activas,
                                       dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                                       lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                                       form_data=form_data) 


        new_instruction = Instruction(
            caminata_id=caminata_id,
            dificultad=dificultad,
            distancia=distancia,
            capacidad=capacidad,
            lugar_salida=lugar_salida,
            fecha_salida=fecha_salida,
            hora_salida=hora_salida,
            fecha_caminata=fecha_caminata,
            hora_inicio_caminata=hora_inicio_caminata,
            recogemos_en=form_data['recogemos_en'], # Usa el JSON ya preparado en form_data
            hidratacion=hidratacion,
            litros_hidratacion=litros_hidratacion,
            tennis_ligera=tennis_ligera,
            tennis_runner=tennis_runner,
            tennis_hiking_baja=tennis_hiking_baja,
            zapato_cana_media=zapato_cana_media,
            zapato_cana_alta=zapato_cana_alta,
            bastones=bastones,
            foco_headlamp=foco_headlamp,
            snacks=snacks,
            repelente=repelente,
            poncho=poncho,
            guantes=guantes,
            bloqueador=bloqueador,
            ropa_cambio=ropa_cambio,
            otras_recomendaciones=otras_recomendaciones,
            normas_generales=normas_generales,
            otras_indicaciones_generales=otras_indicaciones_generales
        )

        try:
            db.session.add(new_instruction)
            db.session.commit()
            flash('Instrucción creada exitosamente!', 'success')
            return redirect(url_for('instrucciones.ver_instrucciones'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear la instrucción: {e}', 'danger')
            current_app.logger.error(f"Error al crear instrucción: {e}")
            return render_template('crear_instrucciones.html', caminatas_activas=caminatas_activas,
                                   dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                                   lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                                   form_data=form_data) 

    return render_template('crear_instrucciones.html', caminatas_activas=caminatas_activas,
                           dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                           lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                           form_data=form_data)

# Ruta para editar una instrucción existente
@instrucciones_bp.route('/editar/<int:instruction_id>', methods=['GET', 'POST'])
def editar_instrucciones(instruction_id):
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))
    
    instruction = Instruction.query.get_or_404(instruction_id)

    caminatas_activas = Caminata.query.filter(Caminata.fecha >= date.today()).order_by(Caminata.fecha).all()
    dificultad_opciones = ["Iniciante", "Básico", "Intermedio", "Avanzado", "Técnico"]
    capacidad_opciones = ["14", "17", "28", "31", "42"]
    lugar_salida_opciones = ["Parque de Tres Ríos - Escuela", "Parque de Tres Ríos - Cruz Roja", "Plaza De San Diego", "Iglesia de San Diego"]
    opciones_sino_opcional = ["SI", "NO", "OPCIONAL"]

    if request.method == 'POST':
        # Actualizar la instrucción con los datos del formulario
        instruction.caminata_id = int(request.form.get('caminata_id'))
        instruction.dificultad = request.form.get('dificultad')
        instruction.distancia = float(request.form.get('distancia')) if request.form.get('distancia') else 0.0
        instruction.capacidad = request.form.get('capacidad')
        instruction.lugar_salida = request.form.get('lugar_salida')
        
        try:
            instruction.fecha_salida = datetime.strptime(request.form.get('fecha_salida'), '%Y-%m-%d').date()
            instruction.hora_salida = datetime.strptime(request.form.get('hora_salida'), '%H:%M').time()
            # Handle optional fecha_caminata and hora_inicio_caminata
            fecha_caminata_str = request.form.get('fecha_caminata')
            hora_inicio_caminata_str = request.form.get('hora_inicio_caminata')
            instruction.fecha_caminata = datetime.strptime(fecha_caminata_str, '%Y-%m-%d').date() if fecha_caminata_str else None
            instruction.hora_inicio_caminata = datetime.strptime(hora_inicio_caminata_str, '%H:%M').time() if hora_inicio_caminata_str else None
        except ValueError as e:
            flash(f'Error en el formato de datos de fecha/hora: {e}', 'danger')
            # Es importante volver a cargar form_data si hay un error para que los campos no se pierdan
            form_data = {
                'caminata_id': request.form.get('caminata_id'),
                'dificultad': request.form.get('dificultad'),
                'distancia': request.form.get('distancia'),
                'capacidad': request.form.get('capacidad'),
                'lugar_salida': request.form.get('lugar_salida'),
                'fecha_salida': request.form.get('fecha_salida'),
                'hora_salida': request.form.get('hora_salida'),
                'fecha_caminata': request.form.get('fecha_caminata'),
                'hora_inicio_caminata': request.form.get('hora_inicio_caminata'),
                'recogemos_en': json.dumps(request.form.getlist('recogemos_en[]')),
                'hidratacion': request.form.get('hidratacion'),
                'litros_hidratacion': request.form.get('litros_hidratacion'),
                'tennis_ligera': request.form.get('tennis_ligera'),
                'tennis_runner': request.form.get('tennis_runner'),
                'tennis_hiking_baja': request.form.get('tennis_hiking_baja'),
                'zapato_cana_media': request.form.get('zapato_cana_media'),
                'zapato_cana_alta': request.form.get('zapato_cana_alta'),
                'bastones': request.form.get('bastones'),
                'foco_headlamp': request.form.get('foco_headlamp'),
                'snacks': request.form.get('snacks'),
                'repelente': request.form.get('repelente'),
                'poncho': request.form.get('poncho'),
                'guantes': request.form.get('guantes'),
                'bloqueador': request.form.get('bloqueador'),
                'ropa_cambio': request.form.get('ropa_cambio'),
                'otras_recomendaciones': request.form.get('otras_recomendaciones'),
                'normas_generales': request.form.get('normas_generales'),
                'otras_indicaciones_generales': request.form.get('otras_indicaciones_generales')
            }
            return render_template('editar_instrucciones.html', instruction=instruction,
                                   caminatas_activas=caminatas_activas, dificultad_opciones=dificultad_opciones,
                                   capacidad_opciones=capacidad_opciones, lugar_salida_opciones=lugar_salida_opciones,
                                   opciones_sino_opcional=opciones_sino_opcional, form_data=form_data)

        recogemos_en_raw = request.form.getlist('recogemos_en[]')
        instruction.recogemos_en = json.dumps([item for item in recogemos_en_raw if item.strip()])


        instruction.hidratacion = request.form.get('hidratacion')
        instruction.litros_hidratacion = request.form.get('litros_hidratacion')
        instruction.tennis_ligera = request.form.get('tennis_ligera')
        instruction.tennis_runner = request.form.get('tennis_runner')
        instruction.tennis_hiking_baja = request.form.get('tennis_hiking_baja')
        instruction.zapato_cana_media = request.form.get('zapato_cana_media')
        instruction.zapato_cana_alta = request.form.get('zapato_cana_alta')
        instruction.bastones = request.form.get('bastones')
        instruction.foco_headlamp = request.form.get('foco_headlamp')
        instruction.snacks = request.form.get('snacks') # Corregido: asignar a instruction.snacks
        instruction.repelente = request.form.get('repelente') # Corregido: asignar a instruction.repelente
        instruction.poncho = request.form.get('poncho') # Corregido: asignar a instruction.poncho
        instruction.guantes = request.form.get('guantes') # Corregido: asignar a instruction.guantes
        instruction.bloqueador = request.form.get('bloqueador') # Corregido: asignar a instruction.bloqueador
        instruction.ropa_cambio = request.form.get('ropa_cambio') # Corregido: asignar a instruction.ropa_cambio


        instruction.otras_recomendaciones = request.form.get('otras_recomendaciones')
        instruction.normas_generales = request.form.get('normas_generales')
        instruction.otras_indicaciones_generales = request.form.get('otras_indicaciones_generales')
        instruction.fecha_modificacion = datetime.now() # Actualizar fecha de modificación


        # Validación de duplicados para fecha y hora de la caminata (excluyendo la instrucción actual)
        # Only validate if both fields are provided
        if instruction.fecha_caminata and instruction.hora_inicio_caminata:
            existing_instruction_query = Instruction.query.filter(
                Instruction.caminata_id == instruction.caminata_id,
                Instruction.fecha_caminata == instruction.fecha_caminata,
                Instruction.hora_inicio_caminata == instruction.hora_inicio_caminata,
                Instruction.id != instruction_id # Excluir la instrucción actual
            )
            if existing_instruction_query.first():
                flash('Ya existe otra instrucción para esta caminata en la fecha y hora de inicio especificadas. Por favor, elige otra hora o fecha para la caminata.', 'danger')
                # Es importante volver a cargar form_data si hay un error para que los campos no se pierdan
                form_data = {
                    'caminata_id': request.form.get('caminata_id'),
                    'dificultad': request.form.get('dificultad'),
                    'distancia': request.form.get('distancia'),
                    'capacidad': request.form.get('capacidad'),
                    'lugar_salida': request.form.get('lugar_salida'),
                    'fecha_salida': request.form.get('fecha_salida'),
                    'hora_salida': request.form.get('hora_salida'),
                    'fecha_caminata': request.form.get('fecha_caminata'),
                    'hora_inicio_caminata': request.form.get('hora_inicio_caminata'),
                    'recogemos_en': json.dumps(request.form.getlist('recogemos_en[]')),
                    'hidratacion': request.form.get('hidratacion'),
                    'litros_hidratacion': request.form.get('litros_hidratacion'),
                    'tennis_ligera': request.form.get('tennis_ligera'),
                    'tennis_runner': request.form.get('tennis_runner'),
                    'tennis_hiking_baja': request.form.get('tennis_hiking_baja'),
                    'zapato_cana_media': request.form.get('zapato_cana_media'),
                    'zapato_cana_alta': request.form.get('zapato_cana_alta'),
                    'bastones': request.form.get('bastones'),
                    'foco_headlamp': request.form.get('foco_headlamp'),
                    'snacks': request.form.get('snacks'),
                    'repelente': request.form.get('repelente'),
                    'poncho': request.form.get('poncho'),
                    'guantes': request.form.get('guantes'),
                    'bloqueador': request.form.get('bloqueador'),
                    'ropa_cambio': request.form.get('ropa_cambio'),
                    'otras_recomendaciones': request.form.get('otras_recomendaciones'),
                    'normas_generales': request.form.get('normas_generales'),
                    'otras_indicaciones_generales': request.form.get('otras_indicaciones_generales')
                }
                return render_template('editar_instrucciones.html', instruction=instruction, caminatas_activas=caminatas_activas,
                                       dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                                       lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                                       form_data=form_data)

        try:
            db.session.commit()
            flash('Instrucción actualizada exitosamente!', 'success')
            return redirect(url_for('instrucciones.detalle_instrucciones', instruction_id=instruction.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la instrucción: {e}', 'danger')
            current_app.logger.error(f"Error al actualizar instrucción {instruction_id}: {e}")
            # Es importante volver a cargar form_data si hay un error para que los campos no se pierdan
            form_data = {
                'caminata_id': request.form.get('caminata_id'),
                'dificultad': request.form.get('dificultad'),
                'distancia': request.form.get('distancia'),
                'capacidad': request.form.get('capacidad'),
                'lugar_salida': request.form.get('lugar_salida'),
                'fecha_salida': request.form.get('fecha_salida'),
                'hora_salida': request.form.get('hora_salida'),
                'fecha_caminata': request.form.get('fecha_caminata'),
                'hora_inicio_caminata': request.form.get('hora_inicio_caminata'),
                'recogemos_en': json.dumps(request.form.getlist('recogemos_en[]')),
                'hidratacion': request.form.get('hidratacion'),
                'litros_hidratacion': request.form.get('litros_hidratacion'),
                'tennis_ligera': request.form.get('tennis_ligera'),
                'tennis_runner': request.form.get('tennis_runner'),
                'tennis_hiking_baja': request.form.get('tennis_hiking_baja'),
                'zapato_cana_media': request.form.get('zapato_cana_media'),
                'zapato_cana_alta': request.form.get('zapato_cana_alta'),
                'bastones': request.form.get('bastones'),
                'foco_headlamp': request.form.get('foco_headlamp'),
                'snacks': request.form.get('snacks'),
                'repelente': request.form.get('repelente'),
                'poncho': request.form.get('poncho'),
                'guantes': request.form.get('guantes'),
                'bloqueador': request.form.get('bloqueador'),
                'ropa_cambio': request.form.get('ropa_cambio'),
                'otras_recomendaciones': request.form.get('otras_recomendaciones'),
                'normas_generales': request.form.get('normas_generales'),
                'otras_indicaciones_generales': request.form.get('otras_indicaciones_generales')
            }
            return render_template('editar_instrucciones.html', instruction=instruction, caminatas_activas=caminatas_activas,
                                   dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                                   lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                                   form_data=form_data)


    # Preparar form_data para mostrar en el formulario GET (o si hay errores)
    form_data = {
        'caminata_id': instruction.caminata_id,
        'dificultad': instruction.dificultad,
        'distancia': instruction.distancia,
        'capacidad': instruction.capacidad,
        'lugar_salida': instruction.lugar_salida,
        'fecha_salida': instruction.fecha_salida.strftime('%Y-%m-%d') if instruction.fecha_salida else '',
        'hora_salida': instruction.hora_salida.strftime('%H:%M') if instruction.hora_salida else '',
        'fecha_caminata': instruction.fecha_caminata.strftime('%Y-%m-%d') if instruction.fecha_caminata else '',
        'hora_inicio_caminata': instruction.hora_inicio_caminata.strftime('%H:%M') if instruction.hora_inicio_caminata else '',
        'recogemos_en': json.loads(instruction.recogemos_en) if instruction.recogemos_en else [],
        'hidratacion': instruction.hidratacion,
        'litros_hidratacion': instruction.litros_hidratacion,
        'tennis_ligera': instruction.tennis_ligera,
        'tennis_runner': instruction.tennis_runner,
        'tennis_hiking_baja': instruction.tennis_hiking_baja,
        'zapato_cana_media': instruction.zapato_cana_media,
        'zapato_cana_alta': instruction.zapato_cana_alta,
        'bastones': instruction.bastones,
        'foco_headlamp': instruction.foco_headlamp,
        'snacks': instruction.snacks,
        'repelente': instruction.repelente,
        'poncho': instruction.poncho,
        'guantes': instruction.guantes,
        'bloqueador': instruction.bloqueador,
        'ropa_cambio': instruction.ropa_cambio,
        'otras_recomendaciones': instruction.otras_recomendaciones,
        'normas_generales': instruction.normas_generales,
        'otras_indicaciones_generales': instruction.otras_indicaciones_generales
    }

    return render_template('editar_instrucciones.html', instruction=instruction, caminatas_activas=caminatas_activas,
                           dificultad_opciones=dificultad_opciones, capacidad_opciones=capacidad_opciones,
                           lugar_salida_opciones=lugar_salida_opciones, opciones_sino_opcional=opciones_sino_opcional,
                           form_data=form_data)


# Ruta para eliminar una instrucción
@instrucciones_bp.route('/eliminar/<int:instruction_id>', methods=['POST'])
def eliminar_instrucciones(instruction_id):
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))
        
    instruction = Instruction.query.get_or_404(instruction_id)
    try:
        db.session.delete(instruction)
        db.session.commit()
        flash('Instrucción eliminada exitosamente!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la instrucción: {e}', 'danger')
        current_app.logger.error(f"Error al eliminar instrucción {instruction_id}: {e}")
    return redirect(url_for('instrucciones.ver_instrucciones'))

# Ruta para ver el detalle de una instrucción
@instrucciones_bp.route('/detalle/<int:instruction_id>')
def detalle_instrucciones(instruction_id):
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))
    
    instruction = Instruction.query.get_or_404(instruction_id)
    
    # Decodificar recogemos_en para mostrar
    recogemos_list = []
    if instruction.recogemos_en:
        try:
            recogemos_list = json.loads(instruction.recogemos_en)
        except json.JSONDecodeError:
            recogemos_list = ["Error al cargar puntos de recogida."]

    return render_template('detalle_instrucciones.html', instruction=instruction, recogemos_list=recogemos_list)

# Ruta para exportar a Excel
@instrucciones_bp.route('/exportar/excel/<int:instruction_id>')
def exportar_excel(instruction_id):
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))

    instruction = Instruction.query.get_or_404(instruction_id) 

    wb = Workbook()
    ws = wb.active
    ws.title = f"Instrucciones {instruction.caminata.nombre}"

    # Encabezado General
    ws.append(["INSTRUCCIONES PARA LA CAMINATA"])
    ws.merge_cells('A1:B1') # Combinar celdas para el título
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
    ws['A1'].alignment = ws['A1'].alignment.copy(horizontal='center')

    # Espacio
    ws.append([]) 

    # Información General de la Caminata
    ws.append(["INFORMACIÓN DE LA CAMINATA"])
    ws.merge_cells('A3:B3')
    ws['A3'].font = ws['A3'].font.copy(bold=True, size=12)

    ws.append(["Caminata:", instruction.caminata.nombre if instruction.caminata else "N/A"])
    # Solo añadir si no es N/A
    if instruction.dificultad and instruction.dificultad != 'N/A':
        ws.append(["Dificultad:", instruction.dificultad])
    if instruction.distancia is not None:
        ws.append(["Distancia:", f"{instruction.distancia} km"])
    if instruction.capacidad and instruction.capacidad != 'N/A':
        ws.append(["Capacidad:", instruction.capacidad])
    if instruction.lugar_salida and instruction.lugar_salida != 'N/A':
        ws.append(["Lugar de Salida:", instruction.lugar_salida])
    if instruction.fecha_salida:
        ws.append(["Fecha de Salida:", instruction.fecha_salida.strftime('%Y-%m-%d')])
    if instruction.hora_salida:
        ws.append(["Hora de Salida:", instruction.hora_salida.strftime('%H:%M')])
    if instruction.fecha_caminata:
        ws.append(["Fecha de Caminata:", instruction.fecha_caminata.strftime('%Y-%m-%d')])
    if instruction.hora_inicio_caminata:
        ws.append(["Hora Inicio Caminata:", instruction.hora_inicio_caminata.strftime('%H:%M')])
    ws.append([])

    # Puntos de Recogida
    ws.append(["RECOGEMOS EN:"])
    ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
    ws[f'A{ws.max_row}'].font = ws[f'A{ws.max_row}'].font.copy(bold=True, size=12)
    recogemos_list = json.loads(instruction.recogemos_en) if instruction.recogemos_en else []
    if recogemos_list:
        for item in recogemos_list:
            ws.append(["-", item])
    else:
        ws.append(["No hay puntos de recogida especificados."])
    ws.append([])

    # Para el Camino
    ws.append(["PARA EL CAMINO:"])
    ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
    ws[f'A{ws.max_row}'].font = ws[f'A{ws.max_row}'].font.copy(bold=True, size=12)

    if instruction.hidratacion and instruction.hidratacion != 'N/A':
        ws.append(["Hidratación:", f"{instruction.hidratacion} - {instruction.litros_hidratacion or ''}"])
    if instruction.tennis_ligera and instruction.tennis_ligera != 'N/A':
        ws.append(["Tennis Ligera:", instruction.tennis_ligera])
    if instruction.tennis_runner and instruction.tennis_runner != 'N/A':
        ws.append(["Tennis Runner:", instruction.tennis_runner])
    if instruction.tennis_hiking_baja and instruction.tennis_hiking_baja != 'N/A':
        ws.append(["Tennis Hiking Baja:", instruction.tennis_hiking_baja])
    if instruction.zapato_cana_media and instruction.zapato_cana_media != 'N/A':
        ws.append(["Zapato Caña Media:", instruction.zapato_cana_media])
    if instruction.zapato_cana_alta and instruction.zapato_cana_alta != 'N/A':
        ws.append(["Zapato Caña Alta:", instruction.zapato_cana_alta])
    if instruction.bastones and instruction.bastones != 'N/A':
        ws.append(["Bastones:", instruction.bastones])
    if instruction.foco_headlamp and instruction.foco_headlamp != 'N/A':
        ws.append(["Foco o Head-lamp:", instruction.foco_headlamp])
    if instruction.snacks and instruction.snacks != 'N/A':
        ws.append(["Snacks:", instruction.snacks])
    if instruction.repelente and instruction.repelente != 'N/A':
        ws.append(["Repelente:", instruction.repelente])
    if instruction.poncho and instruction.poncho != 'N/A':
        ws.append(["Poncho:", instruction.poncho])
    if instruction.guantes and instruction.guantes != 'N/A':
        ws.append(["Guantes:", instruction.guantes])
    if instruction.bloqueador and instruction.bloqueador != 'N/A':
        ws.append(["Bloqueador:", instruction.bloqueador])
    if instruction.ropa_cambio and instruction.ropa_cambio != 'N/A':
        ws.append(["Ropa de Cambio:", instruction.ropa_cambio])
    ws.append([])

    # Otras Recomendaciones (limpiando HTML y añadiendo saltos de línea para guiones)
    otras_recomendaciones_limpias = BeautifulSoup(instruction.otras_recomendaciones or "", 'html.parser').get_text(separator="\n").strip()
    if otras_recomendaciones_limpias and otras_recomendaciones_limpias != 'N/A':
        ws.append(["OTRAS RECOMENDACIONES:"])
        ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
        ws[f'A{ws.max_row}'].font = ws[f'A{ws.max_row}'].font.copy(bold=True, size=12)
        otras_recomendaciones_formateadas = otras_recomendaciones_limpias.replace('-', '\n-').strip() # Añadir salto de línea antes de cada guion
        ws.append([otras_recomendaciones_formateadas])
        ws.append([])

    # Normas Generales (limpiando HTML y añadiendo saltos de línea para guiones)
    normas_generales_limpias = BeautifulSoup(instruction.normas_generales or "", 'html.parser').get_text(separator="\n").strip()
    if normas_generales_limpias and normas_generales_limpias != 'N/A':
        ws.append(["NORMAS GENERALES:"])
        ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
        ws[f'A{ws.max_row}'].font = ws[f'A{ws.max_row}'].font.copy(bold=True, size=12)
        normas_generales_formateadas = normas_generales_limpias.replace('-', '\n-').strip() # Añadir salto de línea antes de cada guion
        ws.append([normas_generales_formateadas])
        ws.append([])

    # Otras Indicaciones Generales (limpiando HTML y añadiendo saltos de línea para guiones)
    otras_indicaciones_generales_limpias = BeautifulSoup(instruction.otras_indicaciones_generales or "", 'html.parser').get_text(separator="\n").strip()
    if otras_indicaciones_generales_limpias and otras_indicaciones_generales_limpias != 'N/A':
        ws.append(["OTRAS INDICACIONES GENERALES:"])
        ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
        ws[f'A{ws.max_row}'].font = ws[f'A{ws.max_row}'].font.copy(bold=True, size=12)
        otras_indicaciones_generales_formateadas = otras_indicaciones_generales_limpias.replace('-', '\n-').strip() # Añadir salto de línea antes de cada guion
        ws.append([otras_indicaciones_generales_formateadas])
        ws.append([])

    # Fechas
    ws.append(["Fecha de Creación:", instruction.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')])
    ws.append(["Última Modificación:", instruction.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S') if instruction.fecha_modificacion else "N/A"])

    # Ajustar ancho de columnas
    # Corrección: Iterar por índice para obtener la letra de la columna de forma segura
    for i in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in ws[column_letter]: # Iterar sobre las celdas de la columna específica
            try:
                # Asegurarse de que no sea una celda fusionada y tenga un valor
                if cell.value is not None and not cell.is_merged and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                # Manejar casos donde cell.value no sea una cadena
                if cell.value is not None and not cell.is_merged:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                # Capturar cualquier otra excepción inesperada
                pass
        
        adjusted_width = (max_length + 2)
        if adjusted_width > 0: # Evitar anchos de columna cero o negativos
            ws.column_dimensions[column_letter].width = adjusted_width


    # Guardar el archivo en un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"instrucciones_{instruction.caminata.nombre.replace(' ', '_').lower()}.xlsx"
    return Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={filename}"})


# Ruta para exportar a PDF (usando ReportLab)
@instrucciones_bp.route('/exportar/pdf/<int:instruction_id>')
def exportar_pdf(instruction_id):
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))
    
    instruction = Instruction.query.get_or_404(instruction_id)

    # Crear un buffer para guardar el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Estilos personalizados
    style_title = styles['h1']
    style_title.alignment = TA_CENTER
    style_heading = styles['h3']
    style_heading.alignment = TA_LEFT
    style_body = styles['Normal']
    style_body.fontSize = 10
    style_small = styles['Normal']
    style_small.fontSize = 8


    # Título
    story.append(Paragraph(f"INSTRUCCIONES PARA LA CAMINATA - {instruction.caminata.nombre.upper() if instruction.caminata else 'N/A'}", style_title))
    story.append(Spacer(1, 0.2 * inch)) # Espacio

    # Información de la Caminata
    story.append(Paragraph("INFORMACIÓN DE LA CAMINATA", style_heading))
    story.append(Paragraph(f"Caminata: {instruction.caminata.nombre if instruction.caminata else 'N/A'}", style_body))
    
    if instruction.dificultad and instruction.dificultad != 'N/A':
        story.append(Paragraph(f"Dificultad: {instruction.dificultad}", style_body))
    if instruction.distancia is not None:
        story.append(Paragraph(f"Distancia: {f'{instruction.distancia} km'}", style_body))
    if instruction.capacidad and instruction.capacidad != 'N/A':
        story.append(Paragraph(f"Capacidad: {instruction.capacidad}", style_body))
    if instruction.lugar_salida and instruction.lugar_salida != 'N/A':
        story.append(Paragraph(f"Lugar de Salida: {instruction.lugar_salida}", style_body))
    if instruction.fecha_salida:
        story.append(Paragraph(f"Fecha de Salida: {instruction.fecha_salida.strftime('%Y-%m-%d')}", style_body))
    if instruction.hora_salida:
        story.append(Paragraph(f"Hora de Salida: {instruction.hora_salida.strftime('%H:%M')}", style_body))
    if instruction.fecha_caminata:
        story.append(Paragraph(f"Fecha de Caminata: {instruction.fecha_caminata.strftime('%Y-%m-%d')}", style_body))
    if instruction.hora_inicio_caminata:
        story.append(Paragraph(f"Hora Inicio Caminata: {instruction.hora_inicio_caminata.strftime('%H:%M')}", style_body))
    story.append(Spacer(1, 0.2 * inch))

    # Puntos de Recogida
    recogemos_list = []
    if instruction.recogemos_en:
        try:
            recogemos_list = json.loads(instruction.recogemos_en)
        except json.JSONDecodeError:
            recogemos_list = ["Error al cargar puntos de recogida."]
    story.append(Paragraph("RECOGEMOS EN:", style_heading))
    if recogemos_list:
        for item in recogemos_list:
            story.append(Paragraph(f"- {item}", style_body))
    else:
        story.append(Paragraph("No hay puntos de recogida especificados.", style_body))
    story.append(Spacer(1, 0.2 * inch))

    # Para el Camino
    story.append(Paragraph("PARA EL CAMINO:", style_heading))
    if instruction.hidratacion and instruction.hidratacion != 'N/A':
        story.append(Paragraph(f"Hidratación: {instruction.hidratacion} - {instruction.litros_hidratacion or ''}", style_body))
    if instruction.tennis_ligera and instruction.tennis_ligera != 'N/A':
        story.append(Paragraph(f"Tennis Ligera: {instruction.tennis_ligera}", style_body))
    if instruction.tennis_runner and instruction.tennis_runner != 'N/A':
        story.append(Paragraph(f"Tennis Runner: {instruction.tennis_runner}", style_body))
    if instruction.tennis_hiking_baja and instruction.tennis_hiking_baja != 'N/A':
        story.append(Paragraph(f"Tennis Hiking Baja: {instruction.tennis_hiking_baja}", style_body))
    if instruction.zapato_cana_media and instruction.zapato_cana_media != 'N/A':
        story.append(Paragraph(f"Zapato Caña Media: {instruction.zapato_cana_media}", style_body))
    if instruction.zapato_cana_alta and instruction.zapato_cana_alta != 'N/A':
        story.append(Paragraph(f"Zapato Caña Alta: {instruction.zapato_cana_alta}", style_body))
    if instruction.bastones and instruction.bastones != 'N/A':
        story.append(Paragraph(f"Bastones: {instruction.bastones}", style_body))
    if instruction.foco_headlamp and instruction.foco_headlamp != 'N/A':
        story.append(Paragraph(f"Foco o Head-lamp: {instruction.foco_headlamp}", style_body))
    if instruction.snacks and instruction.snacks != 'N/A':
        story.append(Paragraph(f"Snacks: {instruction.snacks}", style_body))
    if instruction.repelente and instruction.repelente != 'N/A':
        story.append(Paragraph(f"Repelente: {instruction.repelente}", style_body))
    if instruction.poncho and instruction.poncho != 'N/A':
        story.append(Paragraph(f"Poncho: {instruction.poncho}", style_body))
    if instruction.guantes and instruction.guantes != 'N/A':
        story.append(Paragraph(f"Guantes: {instruction.guantes}", style_body))
    if instruction.bloqueador and instruction.bloqueador != 'N/A':
        story.append(Paragraph(f"Bloqueador: {instruction.bloqueador}", style_body))
    if instruction.ropa_cambio and instruction.ropa_cambio != 'N/A':
        story.append(Paragraph(f"Ropa de Cambio: {instruction.ropa_cambio}", style_body))
    story.append(Spacer(1, 0.2 * inch))

    # Otras Recomendaciones (manejo de contenido HTML como texto plano y añadiendo saltos de línea para guiones)
    otras_recomendaciones_limpias = BeautifulSoup(instruction.otras_recomendaciones or "", 'html.parser').get_text(separator="\n").strip()
    if otras_recomendaciones_limpias and otras_recomendaciones_limpias != 'N/A':
        story.append(Paragraph("OTRAS RECOMENDACIONES:", style_heading))
        otras_recomendaciones_formateadas = otras_recomendaciones_limpias.replace('-', '\n-').strip()
        story.append(Paragraph(otras_recomendaciones_formateadas, style_body))
        story.append(Spacer(1, 0.2 * inch))

    # Normas Generales (manejo de contenido HTML como texto plano y añadiendo saltos de línea para guiones)
    normas_generales_limpias = BeautifulSoup(instruction.normas_generales or "", 'html.parser').get_text(separator="\n").strip()
    if normas_generales_limpias and normas_generales_limpias != 'N/A':
        story.append(Paragraph("NORMAS GENERALES:", style_heading))
        normas_generales_formateadas = normas_generales_limpias.replace('-', '\n-').strip()
        story.append(Paragraph(normas_generales_formateadas, style_body))
        story.append(Spacer(1, 0.2 * inch))

    # Otras Indicaciones Generales (manejo de contenido HTML como texto plano y añadiendo saltos de línea para guiones)
    otras_indicaciones_generales_limpias = BeautifulSoup(instruction.otras_indicaciones_generales or "", 'html.parser').get_text(separator="\n").strip()
    if otras_indicaciones_generales_limpias and otras_indicaciones_generales_limpias != 'N/A':
        story.append(Paragraph("OTRAS INDICACIONES GENERALES:", style_heading))
        otras_indicaciones_generales_formateadas = otras_indicaciones_generales_limpias.replace('-', '\n-').strip()
        story.append(Paragraph(otras_indicaciones_generales_formateadas, style_body))
        story.append(Spacer(1, 0.2 * inch))

    # Fechas de Creación/Modificación
    story.append(Paragraph(f"Fecha de Creación: {instruction.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')}", style_small))
    if instruction.fecha_modificacion:
        story.append(Paragraph(f"Última Modificación: {instruction.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S')}", style_small))

    # Construir el PDF
    doc.build(story)

    # Volver al inicio del buffer para leer su contenido
    buffer.seek(0)

    filename = f"instrucciones_{instruction.caminata.nombre.replace(' ', '_').lower()}.pdf"
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# Ruta para exportar a JPG (usando Pillow)
@instrucciones_bp.route('/exportar/jpg/<int:instruction_id>')
def exportar_jpg(instruction_id):
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))

    instruction = Instruction.query.get_or_404(instruction_id)

    try:
        # Definir dimensiones y color de fondo de la imagen
        img_width = 800
        img_height = 1200 # Aumentado para más contenido
        img = Image.new('RGB', (img_width, img_height), color = (255, 255, 255)) # Blanco
        d = ImageDraw.Draw(img)

        # Cargar fuente (asegúrate de que esta fuente esté disponible en tu sistema o usa ImageFont.load_default())
        try:
            # Puedes probar con una ruta absoluta o una fuente de sistema común
            font_path = "arial.ttf" # O "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf" en Linux, etc.
            font = ImageFont.truetype(font_path, 14) 
            font_bold = ImageFont.truetype(font_path, 16)
            font_title = ImageFont.truetype(font_path, 20)
        except IOError:
            current_app.logger.warning("No se encontró la fuente 'arial.ttf'. Usando la fuente por defecto.")
            font = ImageFont.load_default()
            font_bold = ImageFont.load_default()
            font_title = ImageFont.load_default()

        y_offset = 20 # Margen superior inicial

        # Función para añadir texto y manejar saltos de línea y desbordamiento
        def add_text_to_image(draw_obj, text_content, x, y, font, fill_color, max_width, line_spacing=4):
            words = text_content.split(' ')
            lines = []
            current_line = []
            for word in words:
                test_line = ' '.join(current_line + [word])
                # Calculate bounding box for the test line to check width
                bbox_test_line = draw_obj.textbbox((0, 0), test_line, font=font)
                text_width = bbox_test_line[2] - bbox_test_line[0] # right - left

                if text_width <= max_width:
                    current_line.append(word)
                else:
                    lines.append(' '.join(current_line))
                    current_line = [word]
            lines.append(' '.join(current_line))

            for line in lines:
                draw_obj.text((x, y), line, font=font, fill=fill_color)
                # Calculate bounding box for the actual line being drawn to get its height
                bbox_line = draw_obj.textbbox((0, 0), line, font=font)
                line_height = bbox_line[3] - bbox_line[1] # bottom - top
                y += line_height + line_spacing
                if y > img_height - 20: # Límite inferior para evitar desbordamiento
                    break # Stop drawing if exceeding image height
            return y

        x_start = 50
        max_text_width = img_width - 2 * x_start
        text_color = (0,0,0) # Negro

        # Título
        # Calculate text bounding box for accurate height. Anchor="mm" makes (x,y) the center.
        # We need to calculate height of the rendered text to advance y_offset.
        bbox_title_text = d.textbbox((0, 0), f"INSTRUCCIONES PARA LA CAMINATA - {instruction.caminata.nombre.upper() if instruction.caminata else 'N/A'}", font=font_title)
        title_full_height = bbox_title_text[3] - bbox_title_text[1] # Actual height of the title text
        
        d.text((img_width / 2, y_offset), f"INSTRUCCIONES PARA LA CAMINATA - {instruction.caminata.nombre.upper() if instruction.caminata else 'N/A'}", font=font_title, fill=text_color, anchor="mm")
        y_offset += title_full_height / 2 + 30 # Advance y_offset by half the text height (since it's centered) + buffer.


        # Información General de la Caminata
        bbox_info_heading = d.textbbox((0,0), "INFORMACIÓN DE LA CAMINATA", font=font_bold)
        info_heading_height = bbox_info_heading[3] - bbox_info_heading[1]
        d.text((x_start, y_offset), "INFORMACIÓN DE LA CAMINATA", font=font_bold, fill=text_color)
        y_offset += info_heading_height + 5
        
        y_offset = add_text_to_image(d, f"Caminata: {instruction.caminata.nombre if instruction.caminata else 'N/A'}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.dificultad and instruction.dificultad != 'N/A':
            y_offset = add_text_to_image(d, f"Dificultad: {instruction.dificultad}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.distancia is not None:
            y_offset = add_text_to_image(d, f"Distancia: {f'{instruction.distancia} km'}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.capacidad and instruction.capacidad != 'N/A':
            y_offset = add_text_to_image(d, f"Capacidad: {instruction.capacidad}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.lugar_salida and instruction.lugar_salida != 'N/A':
            y_offset = add_text_to_image(d, f"Lugar de Salida: {instruction.lugar_salida}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.fecha_salida:
            y_offset = add_text_to_image(d, f"Fecha de Salida: {instruction.fecha_salida.strftime('%Y-%m-%d')}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.hora_salida:
            y_offset = add_text_to_image(d, f"Hora de Salida: {instruction.hora_salida.strftime('%H:%M')}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.fecha_caminata:
            y_offset = add_text_to_image(d, f"Fecha de Caminata: {instruction.fecha_caminata.strftime('%Y-%m-%d')}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.hora_inicio_caminata:
            y_offset = add_text_to_image(d, f"Hora Inicio Caminata: {instruction.hora_inicio_caminata.strftime('%H:%M')}", x_start, y_offset, font, text_color, max_text_width)
        y_offset += 20 # Espacio


        # Puntos de Recogida
        recogemos_list = json.loads(instruction.recogemos_en) if instruction.recogemos_en else []
        bbox_recogemos_heading = d.textbbox((0,0), "RECOGEMOS EN:", font=font_bold)
        recogemos_heading_height = bbox_recogemos_heading[3] - bbox_recogemos_heading[1]
        d.text((x_start, y_offset), "RECOGEMOS EN:", font=font_bold, fill=text_color)
        y_offset += recogemos_heading_height + 5
        if recogemos_list:
            for item in recogemos_list:
                y_offset = add_text_to_image(d, f"- {item}", x_start, y_offset, font, text_color, max_text_width)
        else:
            y_offset = add_text_to_image(d, "No hay puntos de recogida especificados.", x_start, y_offset, font, text_color, max_text_width)
        y_offset += 20 # Espacio

        # Para el Camino
        bbox_camino_heading = d.textbbox((0,0), "PARA EL CAMINO:", font=font_bold)
        camino_heading_height = bbox_camino_heading[3] - bbox_camino_heading[1]
        d.text((x_start, y_offset), "PARA EL CAMINO:", font=font_bold, fill=text_color)
        y_offset += camino_heading_height + 5
        
        if instruction.hidratacion and instruction.hidratacion != 'N/A':
            y_offset = add_text_to_image(d, f"Hidratación: {instruction.hidratacion} - {instruction.litros_hidratacion or ''}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.tennis_ligera and instruction.tennis_ligera != 'N/A':
            y_offset = add_text_to_image(d, f"Tennis Ligera: {instruction.tennis_ligera}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.tennis_runner and instruction.tennis_runner != 'N/A':
            y_offset = add_text_to_image(d, f"Tennis Runner: {instruction.tennis_runner}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.tennis_hiking_baja and instruction.tennis_hiking_baja != 'N/A':
            y_offset = add_text_to_image(d, f"Tennis Hiking Baja: {instruction.tennis_hiking_baja}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.zapato_cana_media and instruction.zapato_cana_media != 'N/A':
            y_offset = add_text_to_image(d, f"Zapato Caña Media: {instruction.zapato_cana_media}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.zapato_cana_alta and instruction.zapato_cana_alta != 'N/A':
            y_offset = add_text_to_image(d, f"Zapato Caña Alta: {instruction.zapato_cana_alta}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.bastones and instruction.bastones != 'N/A':
            y_offset = add_text_to_image(d, f"Bastones: {instruction.bastones}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.foco_headlamp and instruction.foco_headlamp != 'N/A':
            y_offset = add_text_to_image(d, f"Foco o Head-lamp: {instruction.foco_headlamp}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.snacks and instruction.snacks != 'N/A':
            y_offset = add_text_to_image(d, f"Snacks: {instruction.snacks}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.repelente and instruction.repelente != 'N/A':
            y_offset = add_text_to_image(d, f"Repelente: {instruction.repelente}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.poncho and instruction.poncho != 'N/A':
            y_offset = add_text_to_image(d, f"Poncho: {instruction.poncho}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.guantes and instruction.guantes != 'N/A':
            y_offset = add_text_to_image(d, f"Guantes: {instruction.guantes}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.bloqueador and instruction.bloqueador != 'N/A':
            y_offset = add_text_to_image(d, f"Bloqueador: {instruction.bloqueador}", x_start, y_offset, font, text_color, max_text_width)
        if instruction.ropa_cambio and instruction.ropa_cambio != 'N/A':
            y_offset = add_text_to_image(d, f"Ropa de Cambio: {instruction.ropa_cambio}", x_start, y_offset, font, text_color, max_text_width)
        y_offset += 20 # Espacio

        # Otras Recomendaciones (limpiando HTML)
        otras_recomendaciones_limpias = BeautifulSoup(instruction.otras_recomendaciones or "", 'html.parser').get_text(separator="\n").strip()
        if otras_recomendaciones_limpias and otras_recomendaciones_limpias != 'N/A':
            bbox_otras_rec_heading = d.textbbox((0,0), "OTRAS RECOMENDACIONES:", font=font_bold)
            otras_rec_heading_height = bbox_otras_rec_heading[3] - bbox_otras_rec_heading[1]
            d.text((x_start, y_offset), "OTRAS RECOMENDACIONES:", font=font_bold, fill=text_color)
            y_offset += otras_rec_heading_height + 5
            otras_recomendaciones_formateadas = otras_recomendaciones_limpias.replace('-', '\n-').strip() # Añadir salto de línea antes de cada guion
            y_offset = add_text_to_image(d, otras_recomendaciones_formateadas, x_start, y_offset, font, text_color, max_text_width)
            y_offset += 20 # Espacio

        # Normas Generales (limpiando HTML)
        normas_generales_limpias = BeautifulSoup(instruction.normas_generales or "", 'html.parser').get_text(separator="\n").strip()
        if normas_generales_limpias and normas_generales_limpias != 'N/A':
            bbox_normas_gen_heading = d.textbbox((0,0), "NORMAS GENERALES:", font=font_bold)
            normas_gen_heading_height = bbox_normas_gen_heading[3] - bbox_normas_gen_heading[1]
            d.text((x_start, y_offset), "NORMAS GENERALES:", font=font_bold, fill=text_color)
            y_offset += normas_gen_heading_height + 5
            normas_generales_formateadas = normas_generales_limpias.replace('-', '\n-').strip() # Añadir salto de línea antes de cada guion
            y_offset = add_text_to_image(d, normas_generales_formateadas, x_start, y_offset, font, text_color, max_text_width)
            y_offset += 20 # Espacio

        # Otras Indicaciones Generales (limpiando HTML)
        otras_indicaciones_generales_limpias = BeautifulSoup(instruction.otras_indicaciones_generales or "", 'html.parser').get_text(separator="\n").strip()
        if otras_indicaciones_generales_limpias and otras_indicaciones_generales_limpias != 'N/A':
            bbox_otras_ind_heading = d.textbbox((0,0), "OTRAS INDICACIONES GENERALES:", font=font_bold)
            otras_ind_heading_height = bbox_otras_ind_heading[3] - bbox_otras_ind_heading[1]
            d.text((x_start, y_offset), "OTRAS INDICACIONES GENERALES:", font=font_bold, fill=text_color)
            y_offset += otras_ind_heading_height + 5
            otras_indicaciones_generales_formateadas = otras_indicaciones_generales_limpias.replace('-', '\n-').strip() # Añadir salto de línea antes de cada guion
            y_offset = add_text_to_image(d, otras_indicaciones_generales_formateadas, x_start, y_offset, font, text_color, max_text_width)
            y_offset += 20 # Espacio

        # Fechas de Creación/Modificación
        bbox_fecha_creacion = d.textbbox((0,0), f"Fecha de Creación: {instruction.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')}", font=font)
        fecha_creacion_height = bbox_fecha_creacion[3] - bbox_fecha_creacion[1]
        d.text((x_start, y_offset), f"Fecha de Creación: {instruction.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')}", font=font, fill=text_color)
        y_offset += fecha_creacion_height + 5

        if instruction.fecha_modificacion:
            bbox_ultima_mod = d.textbbox((0,0), f"Última Modificación: {instruction.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S')}", font=font)
            ultima_mod_height = bbox_ultima_mod[3] - bbox_ultima_mod[1]
            d.text((x_start, y_offset), f"Última Modificación: {instruction.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S')}", font=font, fill=text_color)


        # Guardar la imagen en un buffer
        img_io = io.BytesIO()
        img.save(img_io, 'JPEG', quality=85) # Guardar como JPEG con calidad 85
        img_io.seek(0)

        filename = f"instrucciones_{instruction.caminata.nombre.replace(' ', '_').lower()}.jpg"
        return send_file(img_io, mimetype='image/jpeg', as_attachment=True, download_name=filename)

    except Exception as e:
        flash(f'Error al generar el JPG: {e}', 'danger')
        current_app.logger.error(f"Error al generar JPG para instrucción {instruction_id}: {e}")
        return redirect(url_for('instrucciones.detalle_instrucciones', instruction_id=instruction.id))


# Ruta para exportar a TXT
@instrucciones_bp.route('/exportar/txt/<int:instruction_id>')
def exportar_txt(instruction_id):
    if 'logged_in' not in session or not session['logged_in']:
        flash('Por favor, inicia sesión para acceder a esta página.', 'info')
        return redirect(url_for('login'))

    instruction = Instruction.query.get_or_404(instruction_id)

    recogemos_list = []
    if instruction.recogemos_en:
        try:
            recogemos_list = json.loads(instruction.recogemos_en)
        except json.JSONDecodeError:
            recogemos_list = ["Error al cargar puntos de recogida."]
    
    recogemos_text = "\n".join([f"- {item}" for item in recogemos_list]) if recogemos_list else "No hay puntos de recogida especificados."

    # Limpiar contenido HTML de CKEditor para TXT
    otras_recomendaciones_limpias = BeautifulSoup(instruction.otras_recomendaciones or "", 'html.parser').get_text(separator="\n").strip()
    normas_generales_limpias = BeautifulSoup(instruction.normas_generales or "", 'html.parser').get_text(separator="\n").strip()
    otras_indicaciones_generales_limpias = BeautifulSoup(instruction.otras_indicaciones_generales or "", 'html.parser').get_text(separator="\n").strip()


    content_parts = []
    content_parts.append(f"INSTRUCCIONES PARA LA CAMINATA: {instruction.caminata.nombre if instruction.caminata else 'N/A'}\n\n")
    content_parts.append(f"INFORMACIÓN DE LA CAMINATA:\n")
    content_parts.append(f"  Dificultad: {instruction.dificultad or 'N/A'}\n")
    if instruction.distancia is not None:
        content_parts.append(f"  Distancia: {instruction.distancia} km\n")
    if instruction.capacidad and instruction.capacidad != 'N/A':
        content_parts.append(f"  Capacidad: {instruction.capacidad}\n")
    if instruction.lugar_salida and instruction.lugar_salida != 'N/A':
        content_parts.append(f"  Lugar de Salida: {instruction.lugar_salida}\n")
    if instruction.fecha_salida:
        content_parts.append(f"  Fecha de Salida: {instruction.fecha_salida.strftime('%Y-%m-%d')}\n")
    if instruction.hora_salida:
        content_parts.append(f"  Hora de Salida: {instruction.hora_salida.strftime('%H:%M')}\n")
    if instruction.fecha_caminata:
        content_parts.append(f"  Fecha de Caminata: {instruction.fecha_caminata.strftime('%Y-%m-%d')}\n")
    if instruction.hora_inicio_caminata:
        content_parts.append(f"  Hora Inicio Caminata: {instruction.hora_inicio_caminata.strftime('%H:%M')}\n")
    content_parts.append(f"\nRECOGEMOS EN:\n{recogemos_text}\n\n")

    content_parts.append(f"PARA EL CAMINO:\n")
    if instruction.hidratacion and instruction.hidratacion != 'N/A':
        content_parts.append(f"  Hidratación: {instruction.hidratacion} - {instruction.litros_hidratacion or ''}\n")
    if instruction.tennis_ligera and instruction.tennis_ligera != 'N/A':
        content_parts.append(f"  Tennis Ligera: {instruction.tennis_ligera}\n")
    if instruction.tennis_runner and instruction.tennis_runner != 'N/A':
        content_parts.append(f"  Tennis Runner: {instruction.tennis_runner}\n")
    if instruction.tennis_hiking_baja and instruction.tennis_hiking_baja != 'N/A':
        content_parts.append(f"  Tennis Hiking Baja: {instruction.tennis_hiking_baja}\n")
    if instruction.zapato_cana_media and instruction.zapato_cana_media != 'N/A':
        content_parts.append(f"  Zapato Caña Media: {instruction.zapato_cana_media}\n")
    if instruction.zapato_cana_alta and instruction.zapato_cana_alta != 'N/A':
        content_parts.append(f"  Zapato Caña Alta: {instruction.zapato_cana_alta}\n")
    if instruction.bastones and instruction.bastones != 'N/A':
        content_parts.append(f"  Bastones: {instruction.bastones}\n")
    if instruction.foco_headlamp and instruction.foco_headlamp != 'N/A':
        content_parts.append(f"  Foco o Head-lamp: {instruction.foco_headlamp}\n")
    if instruction.snacks and instruction.snacks != 'N/A':
        content_parts.append(f"  Snacks: {instruction.snacks}\n")
    if instruction.repelente and instruction.repelente != 'N/A':
        content_parts.append(f"  Repelente: {instruction.repelente}\n")
    if instruction.poncho and instruction.poncho != 'N/A':
        content_parts.append(f"  Poncho: {instruction.poncho}\n")
    if instruction.guantes and instruction.guantes != 'N/A':
        content_parts.append(f"  Guantes: {instruction.guantes}\n")
    if instruction.bloqueador and instruction.bloqueador != 'N/A':
        content_parts.append(f"  Bloqueador: {instruction.bloqueador}\n")
    if instruction.ropa_cambio and instruction.ropa_cambio != 'N/A':
        content_parts.append(f"  Ropa de Cambio: {instruction.ropa_cambio}\n")
    content_parts.append(f"\n")

    # Pre-formatear las cadenas con saltos de línea para evitar el error de f-string
    otras_recomendaciones_formateadas = otras_recomendaciones_limpias.replace('-', '\n-').strip()
    normas_generales_formateadas = normas_generales_limpias.replace('-', '\n-').strip()
    otras_indicaciones_generales_formateadas = otras_indicaciones_generales_limpias.replace('-', '\n-').strip()

    if otras_recomendaciones_limpias and otras_recomendaciones_limpias != 'N/A':
        content_parts.append(f"OTRAS RECOMENDACIONES:\n{otras_recomendaciones_formateadas}\n\n")
    
    if normas_generales_limpias and normas_generales_limpias != 'N/A':
        content_parts.append(f"NORMAS GENERALES:\n{normas_generales_formateadas}\n\n")
    
    if otras_indicaciones_generales_limpias and otras_indicaciones_generales_limpias != 'N/A':
        content_parts.append(f"OTRAS INDICACIONES GENERALES:\n{otras_indicaciones_generales_formateadas}\n\n")

    content_parts.append(f"Fecha de Creación: {instruction.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')}\n")
    if instruction.fecha_modificacion:
        content_parts.append(f"Última Modificación: {instruction.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S')}\n")

    content = "".join(content_parts)

    buffer = io.BytesIO()
    buffer.write(content.encode('utf-8'))
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='text/plain',
        as_attachment=True,
        download_name=f'instrucciones_{instruction.caminata.nombre.replace(" ", "_").lower()}.txt'
    )
