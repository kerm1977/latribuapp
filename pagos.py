from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, send_file, Response
from models import db, Pagos, User # Asegúrate de que Pagos y User estén importados
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import io
from openpyxl import Workbook
from fpdf import FPDF
from functools import wraps # Importa wraps para el decorador

# Creación del Blueprint para las rutas de pagos
pagos_bp = Blueprint('pagos', __name__, template_folder='templates')

# Decorador para requerir rol de Superuser
def superuser_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Verifica si el usuario ha iniciado sesión
        if 'logged_in' not in session or not session['logged_in']:
            flash('Por favor, inicia sesión para acceder a esta página.', 'info')
            return redirect(url_for('login'))
        
        user_id = session.get('user_id')
        if not user_id:
            flash('No se pudo verificar el usuario. Por favor, inicia sesión de nuevo.', 'danger')
            return redirect(url_for('login'))

        # Busca el usuario en la base de datos para verificar su rol
        user = User.query.get(user_id)
        if not user or user.role != 'Superuser':
            flash('Acceso denegado. Solo los Superusuarios pueden acceder a esta página.', 'danger')
            return redirect(url_for('login')) # Redirige a la página de login o a una página de inicio
        return f(*args, **kwargs)
    return decorated_function

# Función para verificar si un archivo es una imagen permitida
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif'}

@pagos_bp.route('/')
@superuser_required # Aplica el decorador para restringir el acceso
def ver_pagos():
    """
    Muestra una lista de todos los pagos de caminatas existentes.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    pagos = Pagos.query.order_by(Pagos.fecha_creacion.desc()).all()
    return render_template('ver_pagos.html', pagos=pagos)

@pagos_bp.route('/crear', methods=['GET', 'POST'])
@superuser_required # Aplica el decorador para restringir el acceso
def crear_pagos():
    """
    Permite al usuario crear un nuevo registro de pago de caminata.
    Maneja la lógica de formularios y la carga de imágenes.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    if request.method == 'POST':
        nombre_caminata = request.form.get('nombre_caminata')
        precio_paquete_str = request.form.get('precio_paquete')
        capacidad_str = request.form.get('capacidad')
        tipo_cambio = request.form.get('tipo_cambio')

        cantidad_dias_estadia = request.form.get('cantidad_dias_estadia')
        cantidad_dias_desayuno = request.form.get('cantidad_dias_desayuno')
        cantidad_dias_merienda = request.form.get('cantidad_dias_merienda')
        cantidad_dias_almuerzo = request.form.get('cantidad_dias_almuerzo')
        cantidad_dias_cafe = request.form.get('cantidad_dias_cafe')
        cantidad_dias_cena = request.form.get('cantidad_dias_cena')

        precio_buseta = request.form.get('precio_buseta')
        cantidad_busetas = request.form.get('cantidad_busetas')
        
        precio_acuatico = request.form.get('precio_acuatico')
        cantidad_acuatico = request.form.get('cantidad_acuatico')
        
        precio_aereo = request.form.get('precio_aereo')
        cantidad_aereo = request.form.get('cantidad_aereo')
        
        precio_otros_transporte = request.form.get('precio_otros_transporte')
        cantidad_otros_transporte = request.form.get('cantidad_otros_transporte')
        descripcion_otros_transporte = request.form.get('descripcion_otros_transporte')

        precio_guias_form = request.form.get('precio_guias')
        precio_guia_por_persona_form = request.form.get('precio_guia_por_persona')
        
        cantidad_guias = request.form.get('cantidad_guias')

        precio_estadia = request.form.get('precio_estadia')
        precio_impuestos = request.form.get('precio_impuestos')
        precio_banos = request.form.get('precio_banos')
        precio_servicios_sanitarios = request.form.get('precio_servicios_sanitarios')
        precio_desayuno = request.form.get('precio_desayuno')
        precio_merienda = request.form.get('precio_merienda')
        precio_almuerzo = request.form.get('precio_almuerzo')
        precio_acarreo = request.form.get('precio_acarreo')
        precio_cafe = request.form.get('precio_cafe')
        precio_cena = request.form.get('precio_cena')
        precio_entrada = request.form.get('precio_entrada')
        precio_reconocimiento = request.form.get('precio_reconocimiento')
        precio_permisos = request.form.get('precio_permisos')
        precio_pasaporte = request.form.get('precio_pasaporte')
        
        precio_otros1_personales = request.form.get('precio_otros1_personales')
        descripcion_otros1_personales = request.form.get('descripcion_otros1_personales')
        precio_otros2_personales = request.form.get('precio_otros2_personales')
        descripcion_otros2_personales = request.form.get('descripcion_otros2_personales')
        precio_otros3_personales = request.form.get('precio_otros3_personales')
        descripcion_otros3_personales = request.form.get('descripcion_otros3_personales')
        precio_otros4_personales = request.form.get('precio_otros4_personales')
        descripcion_otros4_personales = request.form.get('descripcion_otros4_personales')

        errors = []
        if not nombre_caminata:
            errors.append('El nombre de la caminata es obligatorio.')
        if not precio_paquete_str:
            errors.append('El precio del paquete es obligatorio.')
        if not capacidad_str:
            errors.append('La capacidad es obligatoria.')

        precio_paquete = 0
        capacidad = 0

        try:
            if precio_paquete_str:
                precio_paquete = int(float(precio_paquete_str))
                if precio_paquete < 0:
                    errors.append('El precio del paquete no puede ser negativo.')
        except ValueError:
            errors.append('El precio del paquete debe ser un número entero válido.')

        try:
            if capacidad_str:
                capacidad = int(float(capacidad_str))
                if capacidad <= 0:
                    errors.append('La capacidad debe ser un número entero mayor que cero.')
        except ValueError:
            errors.append('La capacidad debe ser un número entero válido.')

        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('crear_pagos.html',
                                   nombre_caminata=nombre_caminata,
                                   precio_paquete=precio_paquete_str,
                                   capacidad=capacidad_str,
                                   tipo_cambio=tipo_cambio,
                                   cantidad_dias_estadia=cantidad_dias_estadia,
                                   cantidad_dias_desayuno=cantidad_dias_desayuno,
                                   cantidad_dias_merienda=cantidad_dias_merienda,
                                   cantidad_dias_almuerzo=cantidad_dias_almuerzo,
                                   cantidad_dias_cafe=cantidad_dias_cafe,
                                   cantidad_dias_cena=cantidad_dias_cena,
                                   precio_buseta=precio_buseta,
                                   cantidad_busetas=cantidad_busetas,
                                   precio_acuatico=precio_acuatico,
                                   cantidad_acuatico=cantidad_acuatico,
                                   precio_aereo=precio_aereo,
                                   cantidad_aereo=cantidad_aereo,
                                   precio_otros_transporte=precio_otros_transporte,
                                   cantidad_otros_transporte=cantidad_otros_transporte,
                                   descripcion_otros_transporte=descripcion_otros_transporte,
                                   precio_guias=precio_guias_form,
                                   precio_guia_por_persona=precio_guia_por_persona_form,
                                   cantidad_guias=cantidad_guias,
                                   precio_estadia=precio_estadia,
                                   precio_impuestos=precio_impuestos,
                                   precio_banos=precio_banos,
                                   precio_servicios_sanitarios=precio_servicios_sanitarios,
                                   precio_acarreo=precio_acarreo,
                                   precio_entrada=precio_entrada,
                                   precio_reconocimiento=precio_reconocimiento,
                                   precio_permisos=precio_permisos,
                                   precio_pasaporte=precio_pasaporte,
                                   precio_otros1_personales=precio_otros1_personales,
                                   descripcion_otros1_personales=descripcion_otros1_personales,
                                   precio_otros2_personales=precio_otros2_personales,
                                   descripcion_otros2_personales=descripcion_otros2_personales,
                                   precio_otros3_personales=precio_otros3_personales,
                                   descripcion_otros3_personales=descripcion_otros3_personales,
                                   precio_otros4_personales=precio_otros4_personales,
                                   descripcion_otros4_personales=descripcion_otros4_personales
                                   )

        try:
            tipo_cambio = int(float(tipo_cambio)) if tipo_cambio else 0

            # Se ha cambiado el valor predeterminado de 1 a 0 para las cantidades de días
            cantidad_dias_estadia = int(float(cantidad_dias_estadia)) if cantidad_dias_estadia else 0
            cantidad_dias_desayuno = int(float(cantidad_dias_desayuno)) if cantidad_dias_desayuno else 0
            cantidad_dias_merienda = int(float(cantidad_dias_merienda)) if cantidad_dias_merienda else 0
            cantidad_dias_almuerzo = int(float(cantidad_dias_almuerzo)) if cantidad_dias_almuerzo else 0
            cantidad_dias_cafe = int(float(cantidad_dias_cafe)) if cantidad_dias_cafe else 0
            cantidad_dias_cena = int(float(cantidad_dias_cena)) if cantidad_dias_cena else 0

            precio_buseta = int(float(precio_buseta)) if precio_buseta else 0
            cantidad_busetas = int(float(cantidad_busetas)) if cantidad_busetas else 0
            
            precio_acuatico = int(float(precio_acuatico)) if precio_acuatico else 0
            cantidad_acuatico = int(float(cantidad_acuatico)) if cantidad_acuatico else 0
            
            precio_aereo = int(float(precio_aereo)) if precio_aereo else 0
            cantidad_aereo = int(float(cantidad_aereo)) if cantidad_aereo else 0
            
            precio_otros_transporte = int(float(precio_otros_transporte)) if precio_otros_transporte else 0
            cantidad_otros_transporte = int(float(cantidad_otros_transporte)) if cantidad_otros_transporte else 0

            cantidad_guias = int(float(cantidad_guias)) if cantidad_guias else 0

            precio_estadia = int(float(precio_estadia)) if precio_estadia else 0
            precio_impuestos = int(float(precio_impuestos)) if precio_impuestos else 0
            precio_banos = int(float(precio_banos)) if precio_banos else 0
            precio_servicios_sanitarios = int(float(precio_servicios_sanitarios)) if precio_servicios_sanitarios else 0
            precio_desayuno = int(float(precio_desayuno)) if precio_desayuno else 0
            precio_merienda = int(float(precio_merienda)) if precio_merienda else 0
            precio_almuerzo = int(float(precio_almuerzo)) if precio_almuerzo else 0
            precio_acarreo = int(float(precio_acarreo)) if precio_acarreo else 0
            precio_cafe = int(float(precio_cafe)) if precio_cafe else 0
            precio_cena = int(float(precio_cena)) if precio_cena else 0
            precio_entrada = int(float(precio_entrada)) if precio_entrada else 0
            precio_reconocimiento = int(float(precio_reconocimiento)) if precio_reconocimiento else 0
            precio_permisos = int(float(precio_permisos)) if precio_permisos else 0
            precio_pasaporte = int(float(precio_pasaporte)) if precio_pasaporte else 0
            
            precio_otros1_personales = int(float(precio_otros1_personales)) if precio_otros1_personales else 0
            precio_otros2_personales = int(float(precio_otros2_personales)) if precio_otros2_personales else 0
            precio_otros3_personales = int(float(precio_otros3_personales)) if precio_otros3_personales else 0
            precio_otros4_personales = int(float(precio_otros4_personales)) if precio_otros4_personales else 0

        except ValueError:
            flash('Los campos de precio y cantidad deben ser números enteros válidos.', 'danger')
            return render_template('crear_pagos.html',
                                   nombre_caminata=nombre_caminata,
                                   precio_paquete=precio_paquete_str,
                                   capacidad=capacidad_str,
                                   tipo_cambio=tipo_cambio,
                                   cantidad_dias_estadia=cantidad_dias_estadia,
                                   cantidad_dias_desayuno=cantidad_dias_desayuno,
                                   cantidad_dias_merienda=cantidad_dias_merienda,
                                   cantidad_dias_almuerzo=cantidad_dias_almuerzo,
                                   cantidad_dias_cafe=cantidad_dias_cafe,
                                   cantidad_dias_cena=cantidad_dias_cena,
                                   precio_buseta=precio_buseta,
                                   cantidad_busetas=cantidad_busetas,
                                   precio_acuatico=precio_acuatico,
                                   cantidad_acuatico=cantidad_acuatico,
                                   precio_aereo=precio_aereo,
                                   cantidad_aereo=cantidad_aereo,
                                   precio_otros_transporte=precio_otros_transporte,
                                   cantidad_otros_transporte=cantidad_otros_transporte,
                                   descripcion_otros_transporte=descripcion_otros_transporte,
                                   precio_guias=precio_guias_form,
                                   precio_guia_por_persona=precio_guia_por_persona_form,
                                   cantidad_guias=cantidad_guias,
                                   precio_estadia=precio_estadia,
                                   precio_impuestos=precio_impuestos,
                                   precio_banos=precio_banos,
                                   precio_servicios_sanitarios=precio_servicios_sanitarios,
                                   precio_acarreo=precio_acarreo,
                                   precio_entrada=precio_entrada,
                                   precio_reconocimiento=precio_reconocimiento,
                                   precio_permisos=precio_permisos,
                                   precio_pasaporte=precio_pasaporte,
                                   precio_otros1_personales=precio_otros1_personales,
                                   descripcion_otros1_personales=descripcion_otros1_personales,
                                   precio_otros2_personales=precio_otros2_personales,
                                   descripcion_otros2_personales=descripcion_otros2_personales,
                                   precio_otros3_personales=precio_otros3_personales,
                                   descripcion_otros3_personales=descripcion_otros3_personales,
                                   precio_otros4_personales=precio_otros4_personales,
                                   descripcion_otros4_personales=descripcion_otros4_personales
                                   )


        capacidad_val = capacidad if capacidad > 0 else 1

        total_general_transporte = int((precio_buseta * cantidad_busetas) + \
                                   (precio_acuatico * cantidad_acuatico) + \
                                   (precio_aereo * cantidad_aereo) + \
                                   (precio_otros_transporte * cantidad_otros_transporte))
        total_individual_transporte = int(total_general_transporte / capacidad_val if capacidad_val else 0)

        precio_guias_to_db = 0
        precio_guia_por_persona_to_db = 0
        calculated_total_general_guias = 0
        calculated_total_individual_guias = 0

        parsed_precio_guias_form = float(precio_guias_form) if precio_guias_form else 0
        parsed_precio_guia_por_persona_form = float(precio_guia_por_persona_form) if precio_guia_por_persona_form else 0

        if parsed_precio_guias_form > 0:
            precio_guias_to_db = int(parsed_precio_guias_form)
            calculated_total_general_guias = precio_guias_to_db * cantidad_guias
            calculated_total_individual_guias = int(calculated_total_general_guias / capacidad_val) if capacidad_val else 0
            precio_guia_por_persona_to_db = 0 
        elif parsed_precio_guia_por_persona_form > 0:
            precio_guia_por_persona_to_db = int(parsed_precio_guia_por_persona_form)
            calculated_total_individual_guias = precio_guia_por_persona_to_db * cantidad_guias
            calculated_total_general_guias = int(calculated_total_individual_guias * capacidad_val)
            precio_guias_to_db = 0 
        
        total_general_guias = calculated_total_general_guias
        total_individual_guias = calculated_total_individual_guias

        total_individual_personales = int(
            (precio_estadia * cantidad_dias_estadia) + 
            precio_impuestos + 
            precio_banos + 
            precio_servicios_sanitarios +
            (precio_desayuno * cantidad_dias_desayuno) + 
            (precio_merienda * cantidad_dias_merienda) + 
            (precio_almuerzo * cantidad_dias_almuerzo) + 
            precio_acarreo +
            (precio_cafe * cantidad_dias_cafe) + 
            (precio_cena * cantidad_dias_cena) + 
            precio_entrada + 
            precio_reconocimiento +
            precio_permisos + 
            precio_pasaporte + 
            precio_otros1_personales +
            precio_otros2_personales + 
            precio_otros3_personales + 
            precio_otros4_personales
        )
        
        # ACTUALIZACIÓN DE FÓRMULAS para coincidir con JS
        # total_individual_total = int(total_general_transporte + total_general_guias + total_individual_personales) # OLD
        total_individual_total = int(total_individual_transporte + total_individual_guias + total_individual_personales) # NEW
        total_general_total = int(total_individual_total * capacidad_val)
        ganancia_pp = int(precio_paquete - total_individual_total)
        ganancia_gral = int(ganancia_pp * capacidad)

        flyer_imagen_url = None
        if 'flyer_imagen' in request.files:
            flyer_file = request.files['flyer_imagen']
            if flyer_file and allowed_file(flyer_file.filename):
                filename = secure_filename(flyer_file.filename)
                upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'pagos_images')
                os.makedirs(upload_folder, exist_ok=True)
                file_path = os.path.join(upload_folder, filename)
                flyer_file.save(file_path)
                flyer_imagen_url = 'uploads/pagos_images/' + filename
            elif flyer_file.filename != '':
                flash('Tipo de archivo no permitido. Solo se aceptan PNG, JPG, JPEG, GIF para la imagen.', 'warning')
                return render_template('crear_pagos.html',
                                   nombre_caminata=nombre_caminata,
                                   precio_paquete=precio_paquete_str,
                                   capacidad=capacidad_str,
                                   tipo_cambio=tipo_cambio,
                                   cantidad_dias_estadia=cantidad_dias_estadia,
                                   cantidad_dias_desayuno=cantidad_dias_desayuno,
                                   cantidad_dias_merienda=cantidad_dias_merienda,
                                   cantidad_dias_almuerzo=cantidad_dias_almuerzo,
                                   cantidad_dias_cafe=cantidad_dias_cafe,
                                   cantidad_dias_cena=cantidad_dias_cena,
                                   precio_buseta=precio_buseta,
                                   cantidad_busetas=cantidad_busetas,
                                   precio_acuatico=precio_acuatico,
                                   cantidad_acuatico=cantidad_acuatico,
                                   precio_aereo=precio_aereo,
                                   cantidad_aereo=cantidad_aereo,
                                   precio_otros_transporte=precio_otros_transporte,
                                   cantidad_otros_transporte=cantidad_otros_transporte,
                                   descripcion_otros_transporte=descripcion_otros_transporte,
                                   precio_guias=precio_guias_form,
                                   precio_guia_por_persona=precio_guia_por_persona_form,
                                   cantidad_guias=cantidad_guias,
                                   precio_estadia=precio_estadia,
                                   precio_impuestos=precio_impuestos,
                                   precio_banos=precio_banos,
                                   precio_servicios_sanitarios=precio_servicios_sanitarios,
                                   precio_acarreo=precio_acarreo,
                                   precio_entrada=precio_entrada,
                                   precio_reconocimiento=precio_reconocimiento,
                                   precio_permisos=precio_permisos,
                                   precio_pasaporte=precio_pasaporte,
                                   precio_otros1_personales=precio_otros1_personales,
                                   descripcion_otros1_personales=descripcion_otros1_personales,
                                   precio_otros2_personales=precio_otros2_personales,
                                   descripcion_otros2_personales=descripcion_otros2_personales,
                                   precio_otros3_personales=precio_otros3_personales,
                                   descripcion_otros3_personales=descripcion_otros3_personales,
                                   precio_otros4_personales=precio_otros4_personales,
                                   descripcion_otros4_personales=descripcion_otros4_personales
                                   )

        current_app.logger.debug(f"Datos para Pagos: "
                                 f"user_id={session.get('user_id')}, "
                                 f"nombre_caminata={nombre_caminata}, "
                                 f"precio_paquete={precio_paquete}, "
                                 f"capacidad={capacidad}, "
                                 f"tipo_cambio={tipo_cambio}, "
                                 f"cantidad_dias_estadia={cantidad_dias_estadia}, "
                                 f"cantidad_dias_desayuno={cantidad_dias_desayuno}, "
                                 f"cantidad_dias_merienda={cantidad_dias_merienda}, "
                                 f"cantidad_dias_almuerzo={cantidad_dias_almuerzo}, "
                                 f"cantidad_dias_cafe={cantidad_dias_cafe}, "
                                 f"cantidad_dias_cena={cantidad_dias_cena}, "
                                 f"precio_guias={precio_guias_to_db}, "
                                 f"precio_guia_por_persona={precio_guia_por_persona_to_db}, "
                                 f"total_general_guias={total_general_guias}, "
                                 f"total_individual_guias={total_individual_guias}, "
                                 f"cantidad_guias={cantidad_guias}, "
                                 f"precio_estadia={precio_estadia}, "
                                 f"precio_impuestos={precio_impuestos}, "
                                 f"precio_banos={precio_banos}, "
                                 f"precio_servicios_sanitarios={precio_servicios_sanitarios}, "
                                 f"precio_desayuno={precio_desayuno}, "
                                 f"precio_merienda={precio_merienda}, "
                                 f"precio_almuerzo={precio_almuerzo}, "
                                 f"precio_acarreo={precio_acarreo}, "
                                 f"precio_cafe={precio_cafe}, "
                                 f"precio_cena={precio_cena}, "
                                 f"precio_entrada={precio_entrada}, "
                                 f"precio_reconocimiento={precio_reconocimiento}, "
                                 f"precio_permisos={precio_permisos}, "
                                 f"precio_pasaporte={precio_pasaporte}, "
                                 f"precio_otros1_personales={precio_otros1_personales}, "
                                 f"descripcion_otros1_personales={descripcion_otros1_personales}, "
                                 f"precio_otros2_personales={precio_otros2_personales}, "
                                 f"descripcion_otros2_personales={descripcion_otros2_personales}, "
                                 f"precio_otros3_personales={precio_otros3_personales}, "
                                 f"descripcion_otros3_personales={descripcion_otros3_personales}, "
                                 f"precio_otros4_personales={precio_otros4_personales}, "
                                 f"descripcion_otros4_personales={descripcion_otros4_personales}, "
                                 f"total_individual_personales={total_individual_personales}, "
                                 f"precio_buseta={precio_buseta}, "
                                 f"cantidad_busetas={cantidad_busetas}, "
                                 f"precio_acuatico={precio_acuatico}, "
                                 f"cantidad_acuatico={cantidad_acuatico}, "
                                 f"precio_aereo={precio_aereo}, "
                                 f"cantidad_aereo={cantidad_aereo}, "
                                 f"precio_otros_transporte={precio_otros_transporte}, "
                                 f"cantidad_otros_transporte={cantidad_otros_transporte}, "
                                 f"descripcion_otros_transporte={descripcion_otros_transporte}, "
                                 f"total_general_transporte={total_general_transporte}, "
                                 f"total_individual_transporte={total_individual_transporte}, "
                                 f"total_general_total={total_general_total}, "
                                 f"total_individual_total={total_individual_total}, "
                                 f"ganancia_pp={ganancia_pp}, "
                                 f"ganancia_gral={ganancia_gral}, "
                                 f"flyer_imagen_url={flyer_imagen_url}"
                                )

        new_pago = Pagos(
            user_id=session['user_id'],
            nombre_caminata=nombre_caminata,
            precio_paquete=precio_paquete,
            capacidad=capacidad,
            tipo_cambio=tipo_cambio,
            cantidad_dias_estadia=cantidad_dias_estadia,
            cantidad_dias_desayuno=cantidad_dias_desayuno,
            cantidad_dias_merienda=cantidad_dias_merienda,
            cantidad_dias_almuerzo=cantidad_dias_almuerzo,
            cantidad_dias_cafe=cantidad_dias_cafe,
            cantidad_dias_cena=cantidad_dias_cena,
            precio_buseta=precio_buseta,
            cantidad_busetas=cantidad_busetas,
            precio_acuatico=precio_acuatico,
            cantidad_acuatico=cantidad_acuatico,
            precio_aereo=precio_aereo,
            cantidad_aereo=cantidad_aereo,
            precio_otros_transporte=precio_otros_transporte,
            cantidad_otros_transporte=cantidad_otros_transporte,
            descripcion_otros_transporte=descripcion_otros_transporte,
            precio_guias=precio_guias_to_db,
            precio_guia_por_persona=precio_guia_por_persona_to_db,
            cantidad_guias=cantidad_guias,
            precio_estadia=precio_estadia,
            precio_impuestos=precio_impuestos,
            precio_banos=precio_banos,
            precio_servicios_sanitarios=precio_servicios_sanitarios,
            precio_desayuno=precio_desayuno,
            precio_merienda=precio_merienda,
            precio_almuerzo=precio_almuerzo,
            precio_acarreo=precio_acarreo,
            precio_cafe=precio_cafe,
            precio_cena=precio_cena,
            precio_entrada=precio_entrada,
            precio_reconocimiento=precio_reconocimiento,
            precio_permisos=precio_permisos,
            precio_pasaporte=precio_pasaporte,
            precio_otros1_personales=precio_otros1_personales,
            descripcion_otros1_personales=descripcion_otros1_personales,
            precio_otros2_personales=precio_otros2_personales,
            descripcion_otros2_personales=descripcion_otros2_personales,
            precio_otros3_personales=precio_otros3_personales,
            descripcion_otros3_personales=descripcion_otros3_personales,
            precio_otros4_personales=precio_otros4_personales,
            descripcion_otros4_personales=descripcion_otros4_personales,
            total_individual_personales=total_individual_personales,
            total_general_transporte=total_general_transporte,
            total_individual_transporte=total_individual_transporte,
            total_general_guias=total_general_guias,
            total_individual_guias=total_individual_guias,
            total_general_total=total_general_total,
            total_individual_total=total_individual_total,
            ganancia_pp=ganancia_pp,
            ganancia_gral=ganancia_gral,
            flyer_imagen_url=flyer_imagen_url,
            fecha_creacion=datetime.now()
        )

        try:
            db.session.add(new_pago)
            db.session.commit()
            flash('Pago creado exitosamente!', 'success')
            return redirect(url_for('pagos.ver_pagos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear pago: {e}', 'danger')
            current_app.logger.error(f"Error al crear pago: {e}")
            return render_template('crear_pagos.html',
                                   nombre_caminata=nombre_caminata,
                                   precio_paquete=precio_paquete_str,
                                   capacidad=capacidad_str,
                                   tipo_cambio=tipo_cambio,
                                   cantidad_dias_estadia=cantidad_dias_estadia,
                                   cantidad_dias_desayuno=cantidad_dias_desayuno,
                                   cantidad_dias_merienda=cantidad_dias_merienda,
                                   cantidad_dias_almuerzo=cantidad_dias_almuerzo,
                                   cantidad_dias_cafe=cantidad_dias_cafe,
                                   cantidad_dias_cena=cantidad_dias_cena,
                                   precio_buseta=precio_buseta,
                                   cantidad_busetas=cantidad_busetas,
                                   precio_acuatico=precio_acuatico,
                                   cantidad_acuatico=cantidad_acuatico,
                                   precio_aereo=precio_aereo,
                                   cantidad_aereo=cantidad_aereo,
                                   precio_otros_transporte=precio_otros_transporte,
                                   cantidad_otros_transporte=cantidad_otros_transporte,
                                   descripcion_otros_transporte=descripcion_otros_transporte,
                                   precio_guias=precio_guias_form,
                                   precio_guia_por_persona=precio_guia_por_persona_form,
                                   cantidad_guias=cantidad_guias,
                                   precio_estadia=precio_estadia,
                                   precio_impuestos=precio_impuestos,
                                   precio_banos=precio_banos,
                                   precio_servicios_sanitarios=precio_servicios_sanitarios,
                                   precio_acarreo=precio_acarreo,
                                   precio_entrada=precio_entrada,
                                   precio_reconocimiento=precio_reconocimiento,
                                   precio_permisos=precio_permisos,
                                   precio_pasaporte=precio_pasaporte,
                                   precio_otros1_personales=precio_otros1_personales,
                                   descripcion_otros1_personales=descripcion_otros1_personales,
                                   precio_otros2_personales=precio_otros2_personales,
                                   descripcion_otros2_personales=descripcion_otros2_personales,
                                   precio_otros3_personales=precio_otros3_personales,
                                   descripcion_otros3_personales=descripcion_otros3_personales,
                                   precio_otros4_personales=precio_otros4_personales,
                                   descripcion_otros4_personales=descripcion_otros4_personales
                                   )

    return render_template('crear_pagos.html')

@pagos_bp.route('/detalle/<int:pago_id>')
@superuser_required # Aplica el decorador para restringir el acceso
def detalle_pagos(pago_id):
    """
    Muestra los detalles de un pago de caminata específico.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    pago = Pagos.query.get_or_404(pago_id)

    capacidad_val = int(pago.capacidad) if pago.capacidad and pago.capacidad > 0 else 1

    total_general_busetas = (pago.precio_buseta or 0) * (pago.cantidad_busetas or 0)
    total_individual_busetas = (total_general_busetas / capacidad_val) if capacidad_val else 0

    total_general_acuatico = (pago.precio_acuatico or 0) * (pago.cantidad_acuatico or 0)
    total_individual_acuatico = (total_general_acuatico / capacidad_val) if capacidad_val else 0

    total_general_aereo = (pago.precio_aereo or 0) * (pago.cantidad_aereo or 0)
    total_individual_aereo = (total_general_aereo / capacidad_val) if capacidad_val else 0

    total_general_otros_transporte = (pago.precio_otros_transporte or 0) * (pago.cantidad_otros_transporte or 0)
    total_individual_otros_transporte = (total_general_otros_transporte / capacidad_val) if capacidad_val else 0

    total_general_transporte_display = total_general_busetas + total_general_acuatico + total_general_aereo + total_general_otros_transporte
    total_individual_transporte_display = total_individual_busetas + total_individual_acuatico + total_individual_aereo + total_individual_otros_transporte

    total_general_guias_display = pago.total_general_guias
    total_individual_guias_display = pago.total_individual_guias

    # MODIFICACIÓN: Asegura que la cantidad de días sea 0 si no está definida o es 0
    precio_estadia_calc = (pago.precio_estadia or 0) * (pago.cantidad_dias_estadia or 0)
    precio_desayuno_calc = (pago.precio_desayuno or 0) * (pago.cantidad_dias_desayuno or 0)
    precio_merienda_calc = (pago.precio_merienda or 0) * (pago.cantidad_dias_merienda or 0)
    precio_almuerzo_calc = (pago.precio_almuerzo or 0) * (pago.cantidad_dias_almuerzo or 0)
    precio_cafe_calc = (pago.precio_cafe or 0) * (pago.cantidad_dias_cafe or 0)
    precio_cena_calc = (pago.precio_cena or 0) * (pago.cantidad_dias_cena or 0)

    total_individual_personales_display = (precio_estadia_calc +
                                           (pago.precio_impuestos or 0) +
                                           (pago.precio_banos or 0) +
                                           (pago.precio_servicios_sanitarios or 0) +
                                           precio_desayuno_calc +
                                           precio_merienda_calc +
                                           precio_almuerzo_calc +
                                           (pago.precio_acarreo or 0) +
                                           precio_cafe_calc +
                                           precio_cena_calc +
                                           (pago.precio_entrada or 0) +
                                           (pago.precio_reconocimiento or 0) +
                                           (pago.precio_permisos or 0) +
                                           (pago.precio_pasaporte or 0) +
                                           (pago.precio_otros1_personales or 0) +
                                           (pago.precio_otros2_personales or 0) +
                                           (pago.precio_otros3_personales or 0) +
                                           (pago.precio_otros4_personales or 0))

    # ACTUALIZACIÓN DE FÓRMULAS para coincidir con JS
    # total_individual_total_display = (total_general_transporte_display + total_general_guias_display + total_individual_personales_display) # OLD
    total_individual_total_display = (total_individual_transporte_display + total_individual_guias_display + total_individual_personales_display) # NEW
    total_general_total_display = (total_individual_total_display * capacidad_val)
    ganancia_pp_display = pago.precio_paquete - total_individual_total_display
    ganancia_gral_display = ganancia_pp_display * capacidad_val
    
    # Calcular TOTAL INDIVIDUAL USD
    total_individual_usd = 0
    if (pago.tipo_cambio or 0) > 0:
        total_individual_usd = total_individual_total_display / (pago.tipo_cambio or 1) # Usar 1 para evitar división por cero


    return render_template('detalle_pagos.html', 
                           pago=pago,
                           total_general_transporte_display=total_general_transporte_display,
                           total_individual_transporte_display=total_individual_transporte_display,
                           total_general_guias_display=total_general_guias_display,
                           total_individual_guias_display=total_individual_guias_display,
                           total_individual_personales_display=total_individual_personales_display,
                           total_general_total_display=total_general_total_display,
                           total_individual_total_display=total_individual_total_display,
                           ganancia_pp_display=ganancia_pp_display,
                           ganancia_gral_display=ganancia_gral_display,
                           precio_estadia_calc=precio_estadia_calc,
                           precio_desayuno_calc=precio_desayuno_calc,
                           precio_merienda_calc=precio_merienda_calc,
                           precio_almuerzo_calc=precio_almuerzo_calc,
                           precio_cafe_calc=precio_cafe_calc,
                           precio_cena_calc=precio_cena_calc,
                           total_individual_usd=total_individual_usd # Pasar el nuevo valor a la plantilla
                           )

@pagos_bp.route('/editar/<int:pago_id>', methods=['GET', 'POST'])
@superuser_required # Aplica el decorador para restringir el acceso
def editar_pagos(pago_id):
    """
    Permite al usuario editar un registro de pago de caminata existente.
    Maneja la lógica de formularios, actualización de datos y carga de imágenes.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    pago = Pagos.query.get_or_404(pago_id)

    if request.method == 'POST':
        nombre_caminata = request.form.get('nombre_caminata')
        precio_paquete_str = request.form.get('precio_paquete')
        capacidad_str = request.form.get('capacidad')

        errors = []
        if not nombre_caminata:
            errors.append('El nombre de la caminata es obligatorio.')
        if not precio_paquete_str:
            errors.append('El precio del paquete es obligatorio.')
        if not capacidad_str:
            errors.append('La capacidad es obligatoria.')

        try:
            precio_paquete_val = int(float(precio_paquete_str)) if precio_paquete_str else 0
            if precio_paquete_val < 0:
                errors.append('El precio del paquete no puede ser negativo.')
        except ValueError:
            errors.append('El precio del paquete debe ser un número entero válido.')

        try:
            capacidad_val = int(float(capacidad_str)) if capacidad_str else 0
            if capacidad_val <= 0:
                errors.append('La capacidad debe ser un número entero mayor que cero.')
        except ValueError:
            errors.append('La capacidad debe ser un número entero válido.')

        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('editar_pagos.html', pago=pago)

        pago.nombre_caminata = nombre_caminata
        pago.precio_paquete = precio_paquete_val
        pago.capacidad = capacidad_val
        pago.tipo_cambio = int(float(request.form.get('tipo_cambio'))) if request.form.get('tipo_cambio') else 0
        
        # MODIFICACIÓN: Cambiar el valor predeterminado de 1 a 0 para las cantidades de días
        pago.cantidad_dias_estadia = int(float(request.form.get('cantidad_dias_estadia'))) if request.form.get('cantidad_dias_estadia') else 0
        pago.cantidad_dias_desayuno = int(float(request.form.get('cantidad_dias_desayuno'))) if request.form.get('cantidad_dias_desayuno') else 0
        pago.cantidad_dias_merienda = int(float(request.form.get('cantidad_dias_merienda'))) if request.form.get('cantidad_dias_merienda') else 0
        pago.cantidad_dias_almuerzo = int(float(request.form.get('cantidad_dias_almuerzo'))) if request.form.get('cantidad_dias_almuerzo') else 0
        pago.cantidad_dias_cafe = int(float(request.form.get('cantidad_dias_cafe'))) if request.form.get('cantidad_dias_cafe') else 0
        pago.cantidad_dias_cena = int(float(request.form.get('cantidad_dias_cena'))) if request.form.get('cantidad_dias_cena') else 0


        pago.precio_buseta = int(float(request.form.get('precio_buseta'))) if request.form.get('precio_buseta') else 0
        pago.cantidad_busetas = int(float(request.form.get('cantidad_busetas'))) if request.form.get('cantidad_busetas') else 0
        
        pago.precio_acuatico = int(float(request.form.get('precio_acuatico'))) if request.form.get('precio_acuatico') else 0
        pago.cantidad_acuatico = int(float(request.form.get('cantidad_acuatico'))) if request.form.get('cantidad_acuatico') else 0
        
        pago.precio_aereo = int(float(request.form.get('precio_aereo'))) if request.form.get('precio_aereo') else 0
        pago.cantidad_aereo = int(float(request.form.get('cantidad_aereo'))) if request.form.get('cantidad_aereo') else 0
        
        pago.precio_otros_transporte = int(float(request.form.get('precio_otros_transporte'))) if request.form.get('precio_otros_transporte') else 0
        pago.cantidad_otros_transporte = int(float(request.form.get('cantidad_otros_transporte'))) if request.form.get('cantidad_otros_transporte') else 0
        pago.descripcion_otros_transporte = request.form.get('descripcion_otros_transporte')

        precio_guias_form = request.form.get('precio_guias')
        precio_guia_por_persona_form = request.form.get('precio_guia_por_persona')

        pago.cantidad_guias = int(float(request.form.get('cantidad_guias'))) if request.form.get('cantidad_guias') else 0

        parsed_precio_guias_form = float(precio_guias_form) if precio_guias_form else 0
        parsed_precio_guia_por_persona_form = float(precio_guia_por_persona_form) if precio_guia_por_persona_form else 0

        capacidad_val = pago.capacidad if pago.capacidad > 0 else 1

        if parsed_precio_guias_form > 0:
            pago.precio_guias = int(parsed_precio_guias_form)
            pago.precio_guia_por_persona = 0
            pago.total_general_guias = pago.precio_guias * pago.cantidad_guias
            pago.total_individual_guias = int(pago.total_general_guias / capacidad_val) if capacidad_val else 0
        elif parsed_precio_guia_por_persona_form > 0:
            pago.precio_guias = 0
            pago.precio_guia_por_persona = int(parsed_precio_guia_por_persona_form)
            pago.total_individual_guias = pago.precio_guia_por_persona * pago.cantidad_guias
            pago.total_general_guias = int(pago.total_individual_guias * capacidad_val)
        else:
            pago.precio_guias = 0
            pago.precio_guia_por_persona = 0
            pago.total_general_guias = 0
            pago.total_individual_guias = 0

        pago.precio_estadia = int(float(request.form.get('precio_estadia'))) if request.form.get('precio_estadia') else 0
        pago.precio_impuestos = int(float(request.form.get('precio_impuestos'))) if request.form.get('precio_impuestos') else 0
        pago.precio_banos = int(float(request.form.get('precio_banos'))) if request.form.get('precio_banos') else 0
        pago.precio_servicios_sanitarios = int(float(request.form.get('precio_servicios_sanitarios'))) if request.form.get('precio_servicios_sanitarios') else 0
        pago.precio_desayuno = int(float(request.form.get('precio_desayuno'))) if request.form.get('precio_desayuno') else 0
        pago.precio_merienda = int(float(request.form.get('precio_merienda'))) if request.form.get('precio_merienda') else 0
        pago.precio_almuerzo = int(float(request.form.get('precio_almuerzo'))) if request.form.get('precio_almuerzo') else 0
        pago.precio_acarreo = int(float(request.form.get('precio_acarreo'))) if request.form.get('precio_acarreo') else 0
        pago.precio_cafe = int(float(request.form.get('precio_cafe'))) if request.form.get('precio_cafe') else 0
        pago.precio_cena = int(float(request.form.get('precio_cena'))) if request.form.get('precio_cena') else 0
        pago.precio_entrada = int(float(request.form.get('precio_entrada'))) if request.form.get('precio_entrada') else 0
        pago.precio_reconocimiento = int(float(request.form.get('precio_reconocimiento'))) if request.form.get('precio_reconocimiento') else 0
        pago.precio_permisos = int(float(request.form.get('precio_permisos'))) if request.form.get('precio_permisos') else 0
        pago.precio_pasaporte = int(float(request.form.get('precio_pasaporte'))) if request.form.get('precio_pasaporte') else 0
        pago.precio_otros1_personales = int(float(request.form.get('precio_otros1_personales'))) if request.form.get('precio_otros1_personales') else 0
        pago.descripcion_otros1_personales = request.form.get('descripcion_otros1_personales')
        pago.precio_otros2_personales = int(float(request.form.get('precio_otros2_personales'))) if request.form.get('precio_otros2_personales') else 0
        pago.descripcion_otros2_personales = request.form.get('descripcion_otros2_personales')
        pago.precio_otros3_personales = int(float(request.form.get('precio_otros3_personales'))) if request.form.get('precio_otros3_personales') else 0
        pago.descripcion_otros3_personales = request.form.get('descripcion_otros3_personales')
        pago.precio_otros4_personales = int(float(request.form.get('precio_otros4_personales'))) if request.form.get('precio_otros4_personales') else 0
        pago.descripcion_otros4_personales = request.form.get('descripcion_otros4_personales')

        capacidad_val = pago.capacidad if pago.capacidad > 0 else 1
        pago.total_general_transporte = int((pago.precio_buseta * pago.cantidad_busetas) + \
                                        (pago.precio_acuatico * pago.cantidad_acuatico) + \
                                        (pago.precio_aereo * pago.cantidad_aereo) + \
                                        (pago.precio_otros_transporte * pago.cantidad_otros_transporte))
        pago.total_individual_transporte = int(pago.total_general_transporte / capacidad_val) if capacidad_val else 0

        pago.total_individual_personales = int(
            (pago.precio_estadia * pago.cantidad_dias_estadia) + 
            (pago.precio_impuestos or 0) + 
            (pago.precio_banos or 0) + 
            (pago.precio_servicios_sanitarios or 0) +
            (pago.precio_desayuno * pago.cantidad_dias_desayuno) + 
            (pago.precio_merienda * pago.cantidad_dias_merienda) + 
            (pago.precio_almuerzo * pago.cantidad_dias_almuerzo) + 
            (pago.precio_acarreo or 0) +
            (pago.precio_cafe * pago.cantidad_dias_cafe) + 
            (pago.precio_cena * pago.cantidad_dias_cena) + 
            (pago.precio_entrada or 0) + 
            (pago.precio_reconocimiento or 0) +
            (pago.precio_permisos or 0) + 
            (pago.precio_pasaporte or 0) + 
            (pago.precio_otros1_personales or 0) + 
            (pago.precio_otros2_personales or 0) + 
            (pago.precio_otros3_personales or 0) + 
            (pago.precio_otros4_personales or 0)
        )
        
        # ACTUALIZACIÓN DE FÓRMULAS para coincidir con JS
        # pago.total_individual_total = int(pago.total_general_transporte + pago.total_general_guias + pago.total_individual_personales) # OLD
        pago.total_individual_total = int(pago.total_individual_transporte + pago.total_individual_guias + pago.total_individual_personales) # NEW
        pago.total_general_total = int(pago.total_individual_total * capacidad_val)
        pago.ganancia_pp = int(pago.precio_paquete - pago.total_individual_total)
        pago.ganancia_gral = int(pago.ganancia_pp * pago.capacidad)

        if 'flyer_imagen' in request.files:
            flyer_file = request.files['flyer_imagen']
            if flyer_file and allowed_file(flyer_file.filename):
                if pago.flyer_imagen_url and 'default_caminata.png' not in pago.flyer_imagen_url:
                    try:
                        old_path = os.path.join(current_app.root_path, 'static', pago.flyer_imagen_url)
                        if os.path.exists(old_path):
                            os.remove(old_path)
                    except Exception as e:
                        current_app.logger.warning(f"No se pudo eliminar la imagen antigua: {e}")
                
                filename = secure_filename(flyer_file.filename)
                upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'pagos_images')
                os.makedirs(upload_folder, exist_ok=True)
                file_path = os.path.join(upload_folder, filename)
                flyer_file.save(file_path)
                pago.flyer_imagen_url = os.path.join('uploads', 'pagos_images', filename).replace("\\", "/")
            elif flyer_file.filename != '':
                flash('Tipo de archivo no permitido. Solo se aceptan PNG, JPG, JPEG, GIF para la imagen.', 'warning')
                return render_template('editar_pagos.html', pago=pago)
        
        try:
            db.session.commit()
            flash('Pago actualizado exitosamente!', 'success')
            return redirect(url_for('pagos.detalle_pagos', pago_id=pago.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar pago: {e}', 'danger')
            current_app.logger.error(f"Error al actualizar pago: {e}")
            return render_template('editar_pagos.html', pago=pago)

    return render_template('editar_pagos.html', pago=pago)

@pagos_bp.route('/eliminar/<int:pago_id>', methods=['POST'])
@superuser_required # Aplica el decorador para restringir el acceso
def eliminar_pagos(pago_id):
    """
    Elimina un registro de pago de caminata específico.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    pago = Pagos.query.get_or_404(pago_id)
    
    try:
        if pago.flyer_imagen_url and 'default_caminata.png' not in pago.flyer_imagen_url:
            try:
                file_path = os.path.join(current_app.root_path, 'static', pago.flyer_imagen_url)
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                current_app.logger.warning(f"No se pudo eliminar la imagen del pago {pago.id}: {e}")

        db.session.delete(pago)
        db.session.commit()
        flash('Pago eliminado exitosamente!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar pago: {e}', 'danger')
        current_app.logger.error(f"Error al eliminar pago {pago.id}: {e}")
    
    return redirect(url_for('pagos.ver_pagos'))

@pagos_bp.route('/exportar/excel/<int:pago_id>')
@superuser_required # Aplica el decorador para restringir el acceso
def exportar_excel(pago_id):
    """
    Exporta los detalles de un pago específico a un archivo Excel.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    pago = Pagos.query.get_or_404(pago_id) 

    wb = Workbook()
    ws = wb.active
    ws.title = f"Detalle {pago.nombre_caminata}"

    def format_currency(value):
        # Asegura que el valor sea numérico y lo formatea como entero
        return f"¢{int(value or 0):,}".replace(",", ".")

    capacidad_val = int(pago.capacidad) if pago.capacidad and pago.capacidad > 0 else 1
    
    # MODIFICACIÓN: Asegura que las cantidades de días sean 0 si no están definidas o son 0
    cantidad_dias_estadia_val = int(pago.cantidad_dias_estadia or 0)
    cantidad_dias_desayuno_val = int(pago.cantidad_dias_desayuno or 0)
    cantidad_dias_merienda_val = int(pago.cantidad_dias_merienda or 0)
    cantidad_dias_almuerzo_val = int(pago.cantidad_dias_almuerzo or 0)
    cantidad_dias_cafe_val = int(pago.cantidad_dias_cafe or 0)
    cantidad_dias_cena_val = int(pago.cantidad_dias_cena or 0)


    precio_buseta = int(pago.precio_buseta or 0)
    cantidad_busetas = int(pago.cantidad_busetas or 0)
    total_general_busetas = int(precio_buseta * cantidad_busetas)
    total_individual_busetas = int(total_general_busetas / capacidad_val) if capacidad_val else 0

    precio_acuatico = int(pago.precio_acuatico or 0)
    cantidad_acuatico = int(pago.cantidad_acuatico or 0)
    total_general_acuatico = int(precio_acuatico * cantidad_acuatico)
    total_individual_acuatico = int(total_general_acuatico / capacidad_val) if capacidad_val else 0

    precio_aereo = int(pago.precio_aereo or 0)
    cantidad_aereo = int(pago.cantidad_aereo or 0)
    total_general_aereo = int(precio_aereo * cantidad_aereo)
    total_individual_aereo = int(total_general_aereo / capacidad_val) if capacidad_val else 0

    precio_otros_transporte = int(pago.precio_otros_transporte or 0)
    cantidad_otros_transporte = int(pago.cantidad_otros_transporte or 0)
    total_general_otros_transporte = int(precio_otros_transporte * cantidad_otros_transporte)
    total_individual_otros_transporte = int(total_general_otros_transporte / capacidad_val) if capacidad_val else 0

    total_general_transporte = int(total_general_busetas + total_general_acuatico + total_general_aereo + total_general_otros_transporte)
    total_individual_transporte = int(total_individual_busetas + total_individual_acuatico + total_individual_aereo + total_individual_otros_transporte)

    precio_guias = int(pago.precio_guias or 0)
    precio_guia_por_persona = int(pago.precio_guia_por_persona or 0)
    cantidad_guias = int(pago.cantidad_guias or 0)
    
    total_general_guias = int(pago.total_general_guias or 0)
    total_individual_guias = int(pago.total_individual_guias or 0)

    precio_estadia_calc = (pago.precio_estadia or 0) * cantidad_dias_estadia_val
    precio_desayuno_calc = (pago.precio_desayuno or 0) * cantidad_dias_desayuno_val
    precio_merienda_calc = (pago.precio_merienda or 0) * cantidad_dias_merienda_val
    precio_almuerzo_calc = (pago.precio_almuerzo or 0) * cantidad_dias_almuerzo_val
    precio_cafe_calc = (pago.precio_cafe or 0) * cantidad_dias_cafe_val
    precio_cena_calc = (pago.precio_cena or 0) * cantidad_dias_cena_val

    total_individual_personales = int(precio_estadia_calc + \
                                (pago.precio_impuestos or 0) + \
                                (pago.precio_banos or 0) + \
                                (pago.precio_servicios_sanitarios or 0) + \
                                precio_desayuno_calc + \
                                precio_merienda_calc + \
                                precio_almuerzo_calc + \
                                (pago.precio_acarreo or 0) + \
                                precio_cafe_calc + \
                                precio_cena_calc + \
                                (pago.precio_entrada or 0) + \
                                (pago.precio_reconocimiento or 0) + \
                                (pago.precio_permisos or 0) + \
                                (pago.precio_pasaporte or 0) + \
                                (pago.precio_otros1_personales or 0) + \
                                (pago.precio_otros2_personales or 0) + \
                                (pago.precio_otros3_personales or 0) + \
                                (pago.precio_otros4_personales or 0))
    
    # ACTUALIZACIÓN DE FÓRMULAS para coincidir con JS
    # total_individual_total = int(total_general_transporte + total_general_guias + total_individual_personales) # OLD
    total_individual_total = int(total_individual_transporte + total_individual_guias + total_individual_personales) # NEW
    total_general_total = int(total_individual_total * capacidad_val)
    ganancia_pp = int(pago.precio_paquete - total_individual_total)
    ganancia_gral = int(ganancia_pp * capacidad_val)

    total_individual_usd = 0
    if (pago.tipo_cambio or 0) > 0:
        total_individual_usd = total_individual_total / (pago.tipo_cambio or 1)


    ws.append(["DETALLE DE PAGO - " + pago.nombre_caminata.upper()])
    ws.append([])
    ws.append(["INFORMACIÓN GENERAL"])
    ws.append(["Nombre de la Caminata:", pago.nombre_caminata])
    ws.append(["Precio del Paquete:", format_currency(pago.precio_paquete)])
    ws.append(["Capacidad:", pago.capacidad])
    ws.append(["Tipo de Cambio:", format_currency(pago.tipo_cambio)])
    ws.append([])

    ws.append(["TRANSPORTE"])
    ws.append(["Concepto", "Precio (¢)", "Cantidad", "Total General (¢)", "Total Individual (¢)"])
    
    if precio_buseta > 0 or cantidad_busetas > 0:
        ws.append(["Busetas", format_currency(precio_buseta), cantidad_busetas, format_currency(total_general_busetas), format_currency(total_individual_busetas)])
    if precio_acuatico > 0 or cantidad_acuatico > 0:
        ws.append(["Acuático", format_currency(precio_acuatico), cantidad_acuatico, format_currency(total_general_acuatico), format_currency(total_individual_acuatico)])
    if precio_aereo > 0 or cantidad_aereo > 0:
        ws.append(["Aéreo", format_currency(precio_aereo), cantidad_aereo, format_currency(total_general_aereo), format_currency(total_individual_aereo)])
    
    otros_transporte_desc = pago.descripcion_otros_transporte if pago.descripcion_otros_transporte else ""
    if precio_otros_transporte > 0 or cantidad_otros_transporte > 0 or otros_transporte_desc:
        ws.append([f"Otros {('(' + otros_transporte_desc + ')') if otros_transporte_desc else ''}", format_currency(precio_otros_transporte), cantidad_otros_transporte, format_currency(total_general_otros_transporte), format_currency(total_individual_otros_transporte)])
    
    if total_general_transporte > 0:
        ws.append([])
        ws.append(["Total General Transporte:", "", "", format_currency(total_general_transporte)])
        ws.append(["Total Individual Transporte:", "", "", "", format_currency(total_individual_transporte)])
    ws.append([])

    ws.append(["OTROS GENERALES"])
    ws.append(["Concepto", "Precio (¢)", "Cantidad", "Total General (¢)", "Total Individual (¢)"])
    if precio_guias > 0 or precio_guia_por_persona > 0 or cantidad_guias > 0:
        ws.append(["Guías (por grupo)", format_currency(precio_guias), cantidad_guias, format_currency(total_general_guias), format_currency(total_individual_guias)])
        ws.append(["Guías (por persona - referencia)", format_currency(precio_guia_por_persona), "", "", ""])
    ws.append([])

    ws.append(["OTROS PERSONALES"])
    ws.append(["Concepto", "Monto Base (¢)", "Días", "Monto Total (¢)"])
    
    if pago.precio_estadia > 0 or cantidad_dias_estadia_val > 0: # MODIFICADO
        ws.append(["Estadía", format_currency(pago.precio_estadia), cantidad_dias_estadia_val, format_currency(precio_estadia_calc)])
    if pago.precio_impuestos > 0:
        ws.append(["Impuestos", format_currency(pago.precio_impuestos), "", format_currency(pago.precio_impuestos)])
    if pago.precio_banos > 0:
        ws.append(["Baños", format_currency(pago.precio_banos), "", format_currency(pago.precio_banos)])
    if pago.precio_servicios_sanitarios > 0:
        ws.append(["Servicios Sanitarios", format_currency(pago.precio_servicios_sanitarios), "", format_currency(pago.precio_servicios_sanitarios)])
    if pago.precio_desayuno > 0 or cantidad_dias_desayuno_val > 0: # MODIFICADO
        ws.append(["Desayuno", format_currency(pago.precio_desayuno), cantidad_dias_desayuno_val, format_currency(precio_desayuno_calc)])
    if pago.precio_merienda > 0 or cantidad_dias_merienda_val > 0: # MODIFICADO
        ws.append(["Merienda", format_currency(pago.precio_merienda), cantidad_dias_merienda_val, format_currency(precio_merienda_calc)])
    if pago.precio_almuerzo > 0 or cantidad_dias_almuerzo_val > 0: # MODIFICADO
        ws.append(["Almuerzo", format_currency(pago.precio_almuerzo), cantidad_dias_almuerzo_val, format_currency(precio_almuerzo_calc)])
    if pago.precio_acarreo > 0:
        ws.append(["Acarreo", format_currency(pago.precio_acarreo), "", format_currency(pago.precio_acarreo)])
    if pago.precio_cafe > 0 or cantidad_dias_cafe_val > 0: # MODIFICADO
        ws.append(["Café", format_currency(pago.precio_cafe), cantidad_dias_cafe_val, format_currency(precio_cafe_calc)])
    if pago.precio_cena > 0 or cantidad_dias_cena_val > 0: # MODIFICADO
        ws.append(["Cena", format_currency(pago.precio_cena), cantidad_dias_cena_val, format_currency(precio_cena_calc)])
    if pago.precio_entrada > 0:
        ws.append(["Entrada", format_currency(pago.precio_entrada), "", format_currency(pago.precio_entrada)])
    if pago.precio_reconocimiento > 0:
        ws.append(["Reconocimiento", format_currency(pago.precio_reconocimiento), "", format_currency(pago.precio_reconocimiento)])
    if pago.precio_permisos > 0:
        ws.append(["Permisos", format_currency(pago.precio_permisos), "", format_currency(pago.precio_permisos)])
    if pago.precio_pasaporte > 0:
        ws.append(["Pasaporte", format_currency(pago.precio_pasaporte), "", format_currency(pago.precio_pasaporte)])
    
    if pago.precio_otros1_personales > 0 or pago.descripcion_otros1_personales:
        ws.append([f"Otros 1 ({pago.descripcion_otros1_personales or ''})", format_currency(pago.precio_otros1_personales), "", format_currency(pago.precio_otros1_personales)])
    if pago.precio_otros2_personales > 0 or pago.descripcion_otros2_personales:
        ws.append([f"Otros 2 ({pago.descripcion_otros2_personales or ''})", format_currency(pago.precio_otros2_personales), "", format_currency(pago.precio_otros2_personales)])
    if pago.precio_otros3_personales > 0 or pago.descripcion_otros3_personales:
        ws.append([f"Otros 3 ({pago.descripcion_otros3_personales or ''})", format_currency(pago.precio_otros3_personales), "", format_currency(pago.precio_otros3_personales)])
    if pago.precio_otros4_personales > 0 or pago.descripcion_otros4_personales:
        ws.append([f"Otros 4 ({pago.descripcion_otros4_personales or ''})", format_currency(pago.precio_otros4_personales), "", format_currency(pago.precio_otros4_personales)])
    
    if total_individual_personales > 0:
        ws.append([])
        ws.append(["Total Individual Personales:", "", "", "", format_currency(total_individual_personales)])
    ws.append([])


    ws.append(["TOTALES FINALES"])
    ws.append(["TOTAL GENERAL:", format_currency(total_general_total)])
    ws.append(["TOTAL INDIVIDUAL:", format_currency(total_individual_total)])
    if (pago.tipo_cambio or 0) > 0:
        ws.append(["TOTAL INDIVIDUAL (USD):", f"${total_individual_usd:,.2f}"]) # Formato de USD
    ws.append([])

    ws.append(["GANANCIA"])
    ws.append(["GANANCIA POR PERSONA:", format_currency(ganancia_pp)])
    ws.append(["GANANCIA GENERAL:", format_currency(ganancia_gral)])
    ws.append([])

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    filename = f"{pago.nombre_caminata.replace(' ', '_').lower()}_detalles.xlsx"
    return Response(excel_stream.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": f"attachment;filename={filename}"})
    

@pagos_bp.route('/exportar/pdf/<int:pago_id>')
@superuser_required # Aplica el decorador para restringir el acceso
def exportar_pdf(pago_id):
    """
    Exporta los detalles de un pago específico a un archivo PDF.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    pago = Pagos.query.get_or_404(pago_id)

    capacidad_val = int(pago.capacidad) if pago.capacidad and pago.capacidad > 0 else 1
    
    # MODIFICACIÓN: Asegura que las cantidades de días sean 0 si no están definidas o son 0
    cantidad_dias_estadia_val = int(pago.cantidad_dias_estadia or 0)
    cantidad_dias_desayuno_val = int(pago.cantidad_dias_desayuno or 0)
    cantidad_dias_merienda_val = int(pago.cantidad_dias_merienda or 0)
    cantidad_dias_almuerzo_val = int(pago.cantidad_dias_almuerzo or 0)
    cantidad_dias_cafe_val = int(pago.cantidad_dias_cafe or 0)
    cantidad_dias_cena_val = int(pago.cantidad_dias_cena or 0)

    precio_buseta = int(pago.precio_buseta or 0)
    cantidad_busetas = int(pago.cantidad_busetas or 0)
    total_general_busetas = int(precio_buseta * cantidad_busetas)
    total_individual_busetas = int(total_general_busetas / capacidad_val) if capacidad_val else 0

    precio_acuatico = int(pago.precio_acuatico or 0)
    cantidad_acuatico = int(pago.cantidad_acuatico or 0)
    total_general_acuatico = int(precio_acuatico * cantidad_acuatico)
    total_individual_acuatico = int(total_general_acuatico / capacidad_val) if capacidad_val else 0

    precio_aereo = int(pago.precio_aereo or 0)
    cantidad_aereo = int(pago.cantidad_aereo or 0)
    total_general_aereo = int(precio_aereo * cantidad_aereo)
    total_individual_aereo = int(total_general_aereo / capacidad_val) if capacidad_val else 0

    precio_otros_transporte = int(pago.precio_otros_transporte or 0)
    cantidad_otros_transporte = int(pago.cantidad_otros_transporte or 0)
    total_general_otros_transporte = int(precio_otros_transporte * cantidad_otros_transporte)
    total_individual_otros_transporte = int(total_general_otros_transporte / capacidad_val) if capacidad_val else 0

    total_general_transporte = int(total_general_busetas + total_general_acuatico + total_general_aereo + total_general_otros_transporte)
    total_individual_transporte = int(total_individual_busetas + total_individual_acuatico + total_individual_aereo + total_individual_otros_transporte)

    precio_guias = int(pago.precio_guias or 0)
    precio_guia_por_persona = int(pago.precio_guia_por_persona or 0)
    cantidad_guias = int(pago.cantidad_guias or 0)

    total_general_guias = int(pago.total_general_guias or 0)
    total_individual_guias = int(pago.total_individual_guias or 0)

    precio_estadia_calc = (pago.precio_estadia or 0) * cantidad_dias_estadia_val
    precio_desayuno_calc = (pago.precio_desayuno or 0) * cantidad_dias_desayuno_val
    precio_merienda_calc = (pago.precio_merienda or 0) * cantidad_dias_merienda_val
    precio_almuerzo_calc = (pago.precio_almuerzo or 0) * cantidad_dias_almuerzo_val
    precio_cafe_calc = (pago.precio_cafe or 0) * cantidad_dias_cafe_val
    precio_cena_calc = (pago.precio_cena or 0) * cantidad_dias_cena_val

    total_individual_personales = int(precio_estadia_calc + \
                                 (pago.precio_impuestos or 0) + \
                                 (pago.precio_banos or 0) + \
                                 (pago.precio_servicios_sanitarios or 0) + \
                                 precio_desayuno_calc + \
                                 precio_merienda_calc + \
                                 precio_almuerzo_calc + \
                                 (pago.precio_acarreo or 0) + \
                                 precio_cafe_calc + \
                                 precio_cena_calc + \
                                 (pago.precio_entrada or 0) + \
                                 (pago.precio_reconocimiento or 0) + \
                                 (pago.precio_permisos or 0) + \
                                 (pago.precio_pasaporte or 0) + \
                                 (pago.precio_otros1_personales or 0) + \
                                 (pago.precio_otros2_personales or 0) + \
                                 (pago.precio_otros3_personales or 0) + \
                                 (pago.precio_otros4_personales or 0))
    
    # ACTUALIZACIÓN DE FÓRMULAS para coincidir con JS
    # total_individual_total = int(total_general_transporte + total_general_guias + total_individual_personales) # OLD
    total_individual_total = int(total_individual_transporte + total_individual_guias + total_individual_personales) # NEW
    total_general_total = int(total_individual_total * capacidad_val)
    ganancia_pp = int(pago.precio_paquete - total_individual_total)
    ganancia_gral = int(ganancia_pp * capacidad_val)

    total_individual_usd = 0
    if (pago.tipo_cambio or 0) > 0:
        total_individual_usd = total_individual_total / (pago.tipo_cambio or 1)


    class PDF(FPDF):
        def header(self):
            self.set_font('Arial', 'B', 12)
            self.cell(0, 10, 'Detalle de Pago de Caminata', 0, 1, 'C')
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font('Arial', 'I', 8)
            self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', 0, 0, 'C')

        def chapter_title(self, title):
            self.set_font('Arial', 'B', 10)
            self.set_fill_color(200, 220, 255)
            self.cell(0, 8, title, 0, 1, 'L', 1)
            self.ln(2)

        def chapter_body(self, label, value, condition=True): # Añadido parámetro condition
            if condition: # Solo imprime si la condición es True
                self.set_font('Arial', '', 9)
                self.multi_cell(0, 5, f'{label}: {value}', 0, 'L')
                self.ln(1)
        
    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.chapter_title("INFORMACIÓN GENERAL")
    pdf.chapter_body("Nombre de la Caminata", pago.nombre_caminata)
    pdf.chapter_body("Precio del Paquete", f"¢{int(pago.precio_paquete or 0):,.0f}")
    pdf.chapter_body("Capacidad", str(int(pago.capacidad or 0)))
    pdf.chapter_body("Tipo de Cambio", f"¢{int(pago.tipo_cambio or 0):,.0f}", condition=(pago.tipo_cambio or 0) > 0)

    # MODIFICACIÓN: La condición ahora verifica si el valor es mayor que 0
    pdf.chapter_body("Cantidad Días Estadía", str(int(cantidad_dias_estadia_val)), condition=cantidad_dias_estadia_val > 0)
    pdf.chapter_body("Cantidad Días Desayuno", str(int(cantidad_dias_desayuno_val)), condition=cantidad_dias_desayuno_val > 0)
    pdf.chapter_body("Cantidad Días Merienda", str(int(cantidad_dias_merienda_val)), condition=cantidad_dias_merienda_val > 0)
    pdf.chapter_body("Cantidad Días Almuerzo", str(int(cantidad_dias_almuerzo_val)), condition=cantidad_dias_almuerzo_val > 0)
    pdf.chapter_body("Cantidad Días Café", str(int(cantidad_dias_cafe_val)), condition=cantidad_dias_cafe_val > 0)
    pdf.chapter_body("Cantidad Días Cena", str(int(cantidad_dias_cena_val)), condition=cantidad_dias_cena_val > 0)


    pdf.chapter_title("TRANSPORTE")
    pdf.chapter_body("Costo Buseta", f"¢{int(precio_buseta):,.0f} (x{int(cantidad_busetas)})", condition=precio_buseta > 0 or cantidad_busetas > 0)
    pdf.chapter_body("Costo Acuático", f"¢{int(precio_acuatico):,.0f} (x{int(cantidad_acuatico)})", condition=precio_acuatico > 0 or cantidad_acuatico > 0)
    pdf.chapter_body("Costo Aéreo", f"¢{int(precio_aereo):,.0f} (x{int(cantidad_aereo)})", condition=precio_aereo > 0 or cantidad_aereo > 0)
    
    otros_transporte_desc_pdf = pago.descripcion_otros_transporte if pago.descripcion_otros_transporte else ""
    pdf.chapter_body("Costo Otros Transporte", f"¢{int(precio_otros_transporte):,.0f} (x{int(cantidad_otros_transporte)}) {('(' + otros_transporte_desc_pdf + ')') if otros_transporte_desc_pdf else ''}", condition=precio_otros_transporte > 0 or cantidad_otros_transporte > 0 or otros_transporte_desc_pdf)
    
    if total_general_transporte > 0:
        pdf.chapter_body("Total General Transporte", f"¢{int(total_general_transporte):,.0f}")
        pdf.chapter_body("Total Individual Transporte", f"¢{int(total_individual_transporte):,.0f}")

    pdf.chapter_title("GUÍAS")
    pdf.chapter_body("Precio Guías (por grupo)", f"¢{int(precio_guias):,.0f}", condition=precio_guias > 0)
    pdf.chapter_body("Precio Guías (por persona - referencia)", f"¢{int(precio_guia_por_persona):,.0f}", condition=precio_guia_por_persona > 0)
    pdf.chapter_body("Cantidad Guías", f"{int(cantidad_guias)}", condition=cantidad_guias > 0)
    if total_general_guias > 0:
        pdf.chapter_body("Total General Guías", f"¢{int(total_general_guias):,.0f}")
        pdf.chapter_body("Total Individual Guías", f"¢{int(total_individual_guias):,.0f}")

    pdf.chapter_title("GASTOS PERSONALES")
    # MODIFICACIÓN: La condición ahora verifica si el valor base o la cantidad de días es mayor que 0
    pdf.chapter_body("Precio Estadía (base)", f"¢{int(pago.precio_estadia or 0):,.0f} (x{int(cantidad_dias_estadia_val)} días)", condition=(pago.precio_estadia or 0) > 0 or cantidad_dias_estadia_val > 0)
    pdf.chapter_body("Precio Estadía (total)", f"¢{int(precio_estadia_calc):,.0f}", condition=precio_estadia_calc > 0)
    pdf.chapter_body("Precio Impuestos", f"¢{int(pago.precio_impuestos or 0):,.0f}", condition=(pago.precio_impuestos or 0) > 0)
    pdf.chapter_body("Precio Baños", f"¢{int(pago.precio_banos or 0):,.0f}", condition=(pago.precio_banos or 0) > 0)
    pdf.chapter_body("Precio Servicios Sanitarios", f"¢{int(pago.precio_servicios_sanitarios or 0):,.0f}", condition=(pago.precio_servicios_sanitarios or 0) > 0)
    pdf.chapter_body("Precio Desayuno (base)", f"¢{int(pago.precio_desayuno or 0):,.0f} (x{int(cantidad_dias_desayuno_val)} días)", condition=(pago.precio_desayuno or 0) > 0 or cantidad_dias_desayuno_val > 0) # MODIFICADO
    pdf.chapter_body("Precio Desayuno (total)", f"¢{int(precio_desayuno_calc):,.0f}", condition=precio_desayuno_calc > 0)
    pdf.chapter_body("Precio Merienda (base)", f"¢{int(pago.precio_merienda or 0):,.0f} (x{int(cantidad_dias_merienda_val)} días)", condition=(pago.precio_merienda or 0) > 0 or cantidad_dias_merienda_val > 0) # MODIFICADO
    pdf.chapter_body("Precio Merienda (total)", f"¢{int(precio_merienda_calc):,.0f}", condition=precio_merienda_calc > 0)
    pdf.chapter_body("Precio Almuerzo (base)", f"¢{int(pago.precio_almuerzo or 0):,.0f} (x{int(cantidad_dias_almuerzo_val)} días)", condition=(pago.precio_almuerzo or 0) > 0 or cantidad_dias_almuerzo_val > 0) # MODIFICADO
    pdf.chapter_body("Precio Almuerzo (total)", f"¢{int(precio_almuerzo_calc):,.0f}", condition=precio_almuerzo_calc > 0)
    pdf.chapter_body("Precio Acarreo", f"¢{int(pago.precio_acarreo or 0):,.0f}", condition=(pago.precio_acarreo or 0) > 0)
    pdf.chapter_body("Precio Café (base)", f"¢{int(pago.precio_cafe or 0):,.0f} (x{int(cantidad_dias_cafe_val)} días)", condition=(pago.precio_cafe or 0) > 0 or cantidad_dias_cafe_val > 0) # MODIFICADO
    pdf.chapter_body("Precio Café (total)", f"¢{int(precio_cafe_calc):,.0f}", condition=precio_cafe_calc > 0)
    pdf.chapter_body("Precio Cena (base)", f"¢{int(pago.precio_cena or 0):,.0f} (x{int(cantidad_dias_cena_val)} días)", condition=(pago.precio_cena or 0) > 0 or cantidad_dias_cena_val > 0) # MODIFICADO
    pdf.chapter_body("Precio Cena (total)", f"¢{int(precio_cena_calc):,.0f}", condition=precio_cena_calc > 0)
    pdf.chapter_body("Precio Entrada", f"¢{int(pago.precio_entrada or 0):,.0f}", condition=(pago.precio_entrada or 0) > 0)
    pdf.chapter_body("Precio Reconocimiento", f"¢{int(pago.precio_reconocimiento or 0):,.0f}", condition=(pago.precio_reconocimiento or 0) > 0)
    pdf.chapter_body("Precio Permisos", f"¢{int(pago.precio_permisos or 0):,.0f}", condition=(pago.precio_permisos or 0) > 0)
    pdf.chapter_body("Precio Pasaporte", f"¢{int(pago.precio_pasaporte or 0):,.0f}", condition=(pago.precio_pasaporte or 0) > 0)
    
    if pago.precio_otros1_personales > 0 or pago.descripcion_otros1_personales:
        pdf.chapter_body("Precio Otros1", f"¢{int(pago.precio_otros1_personales or 0):,.0f} ({pago.descripcion_otros1_personales or ''})")
    if pago.precio_otros2_personales > 0 or pago.descripcion_otros2_personales:
        pdf.chapter_body("Precio Otros2", f"¢{int(pago.precio_otros2_personales or 0):,.0f} ({pago.descripcion_otros2_personales or ''})")
    if pago.precio_otros3_personales > 0 or pago.descripcion_otros3_personales:
        pdf.chapter_body("Precio Otros3", f"¢{int(pago.precio_otros3_personales or 0):,.0f} ({pago.descripcion_otros3_personales or ''})")
    if pago.precio_otros4_personales > 0 or pago.descripcion_otros4_personales:
        pdf.chapter_body("Precio Otros4", f"¢{int(pago.precio_otros4_personales or 0):,.0f} ({pago.descripcion_otros4_personales or ''})")
    
    if total_individual_personales > 0:
        pdf.chapter_body("Total Individual Personales", f"¢{int(total_individual_personales):,.0f}") 

    pdf.chapter_title("TOTALES FINALES")
    pdf.chapter_body("TOTAL GENERAL", f"¢{int(total_general_total):,.0f}") 
    pdf.chapter_body("TOTAL INDIVIDUAL", f"¢{int(total_individual_total):,.0f}") 
    if (pago.tipo_cambio or 0) > 0:
        pdf.chapter_body("TOTAL INDIVIDUAL (USD)", f"${total_individual_usd:,.2f}") # Formato de USD
    pdf.chapter_body("GANANCIA POR PERSONA", f"¢{int(ganancia_pp):,.0f}")
    pdf.chapter_body("GANANCIA GENERAL", f"¢{int(ganancia_gral):,.0f}")
    pdf.ln(10)

    pdf_output = pdf.output(dest='S').encode('latin1')
    
    return Response(pdf_output, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment;filename=detalle_pago_{pago.nombre_caminata}.pdf'})


@pagos_bp.route('/exportar/txt/<int:pago_id>')
@superuser_required # Aplica el decorador para restringir el acceso
def exportar_txt(pago_id):
    """
    Exporta los detalles de un pago específico a un archivo de texto plano.
    Requiere que el usuario esté logueado y sea Superusuario.
    """
    pago = Pagos.query.get_or_404(pago_id)

    capacidad_val = int(pago.capacidad) if pago.capacidad and pago.capacidad > 0 else 1
    
    # MODIFICACIÓN: Asegura que las cantidades de días sean 0 si no están definidas o son 0
    cantidad_dias_estadia_val = int(pago.cantidad_dias_estadia or 0)
    cantidad_dias_desayuno_val = int(pago.cantidad_dias_desayuno or 0)
    cantidad_dias_merienda_val = int(pago.cantidad_dias_merienda or 0)
    cantidad_dias_almuerzo_val = int(pago.cantidad_dias_almuerzo or 0)
    cantidad_dias_cafe_val = int(pago.cantidad_dias_cafe or 0)
    cantidad_dias_cena_val = int(pago.cantidad_dias_cena or 0)


    total_general_busetas = int((pago.precio_buseta or 0) * (pago.cantidad_busetas or 0))
    total_individual_busetas = int(total_general_busetas / capacidad_val) if capacidad_val else 0

    total_general_acuatico = int((pago.precio_acuatico or 0) * (pago.cantidad_acuatico or 0))
    total_individual_acuatico = int(total_general_acuatico / capacidad_val) if capacidad_val else 0

    total_general_aereo = int((pago.precio_aereo or 0) * (pago.cantidad_aereo or 0))
    total_individual_aereo = int(total_general_aereo / capacidad_val) if capacidad_val else 0

    total_general_otros_transporte = int((pago.precio_otros_transporte or 0) * (pago.cantidad_otros_transporte or 0))
    total_individual_otros_transporte = int(total_general_otros_transporte / capacidad_val) if capacidad_val else 0

    total_general_transporte = int(total_general_busetas + total_general_acuatico + total_general_aereo + total_general_otros_transporte)
    total_individual_transporte = int(total_individual_busetas + total_individual_acuatico + total_individual_aereo + total_individual_otros_transporte)

    total_general_guias = int(pago.total_general_guias or 0)
    total_individual_guias = int(pago.total_individual_guias or 0)

    precio_estadia_calc_display = (pago.precio_estadia or 0) * cantidad_dias_estadia_val
    precio_desayuno_calc_display = (pago.precio_desayuno or 0) * cantidad_dias_desayuno_val
    precio_merienda_calc_display = (pago.precio_merienda or 0) * cantidad_dias_merienda_val
    precio_almuerzo_calc_display = (pago.precio_almuerzo or 0) * cantidad_dias_almuerzo_val
    precio_cafe_calc_display = (pago.precio_cafe or 0) * cantidad_dias_cafe_val
    precio_cena_calc_display = (pago.precio_cena or 0) * cantidad_dias_cena_val

    total_individual_personales_calc = int(precio_estadia_calc_display + \
                                 (pago.precio_impuestos or 0) + \
                                 (pago.precio_banos or 0) + \
                                 (pago.precio_servicios_sanitarios or 0) + \
                                 precio_desayuno_calc_display + \
                                 precio_merienda_calc_display + \
                                 precio_almuerzo_calc_display + \
                                 (pago.precio_acarreo or 0) + \
                                 precio_cafe_calc_display + \
                                 precio_cena_calc_display + \
                                 (pago.precio_entrada or 0) + \
                                 (pago.precio_reconocimiento or 0) + \
                                 (pago.precio_permisos or 0) + \
                                 (pago.precio_pasaporte or 0) + \
                                 (pago.precio_otros1_personales or 0) + \
                                 (pago.precio_otros2_personales or 0) + \
                                 (pago.precio_otros3_personales or 0) + \
                                 (pago.precio_otros4_personales or 0))
    
    total_individual_personales_display = total_individual_personales_calc 

    # ACTUALIZACIÓN DE FÓRMULAS para coincidir con JS
    # total_individual_total_calc = int(total_general_transporte + total_general_guias + total_individual_personales_calc) # OLD
    total_individual_total_calc = int(total_individual_transporte + total_individual_guias + total_individual_personales_calc) # NEW
    total_general_total_calc = int(total_individual_total_calc * capacidad_val)
    ganancia_pp = int((pago.precio_paquete or 0) - total_individual_total_calc)
    ganancia_gral = int(ganancia_pp * (pago.capacidad or 0))

    total_individual_usd = 0
    if (pago.tipo_cambio or 0) > 0:
        total_individual_usd = total_individual_total_calc / (pago.tipo_cambio or 1)


    content = f"Detalles del Pago de Caminata: {pago.nombre_caminata}\n\n" \
              f"Información:\n" \
              f"  Nombre de la caminata: {pago.nombre_caminata}\n" \
              f"  Precio del Paquete: ¢{int(pago.precio_paquete or 0):,.0f}\n" \
              f"  Capacidad: {int(pago.capacidad or 0)}\n" 
    
    if (pago.tipo_cambio or 0) > 0:
        content += f"  Tipo de Cambio: ¢{int(pago.tipo_cambio or 0):,.0f}\n" 
    if cantidad_dias_estadia_val > 0:
        content += f"  Cant. Días Estadía: {int(cantidad_dias_estadia_val)}\n" 
    if cantidad_dias_desayuno_val > 0:
        content += f"  Cant. Días Desayuno: {int(cantidad_dias_desayuno_val)}\n" 
    if cantidad_dias_merienda_val > 0:
        content += f"  Cant. Días Merienda: {int(cantidad_dias_merienda_val)}\n" 
    if cantidad_dias_almuerzo_val > 0:
        content += f"  Cant. Días Almuerzo: {int(cantidad_dias_almuerzo_val)}\n" 
    if cantidad_dias_cafe_val > 0:
        content += f"  Cant. Días Café: {int(cantidad_dias_cafe_val)}\n" 
    if cantidad_dias_cena_val > 0:
        content += f"  Cant. Días Cena: {int(cantidad_dias_cena_val)}\n\n" 
    
    content += f"Transporte:\n" 
    if (pago.precio_buseta or 0) > 0 or (pago.cantidad_busetas or 0) > 0: # Corrected line
        content += f"  Precio Buseta: ¢{int(pago.precio_buseta or 0):,.0f}\n" \
                   f"  Cantidad Busetas: {int(pago.cantidad_busetas or 0)}\n" \
                   f"  Total General Busetas: ¢{int(total_general_busetas):,.0f}\n" \
                   f"  Total Individual Busetas: ¢{int(total_individual_busetas):,.0f}\n\n" 
    
    if (pago.precio_acuatico or 0) > 0 or (pago.cantidad_acuatico or 0) > 0: # Corrected line
        content += f"  Precio Acuatico: ¢{int(pago.precio_acuatico or 0):,.0f}\n" \
                   f"  Cantidad Acuatico: {int(pago.cantidad_acuatico or 0)}\n" \
                   f"  Total General Acuatico: ¢{int(total_general_acuatico):,.0f}\n" \
                   f"  Total Individual Acuatico: ¢{int(total_individual_acuatico):,.0f}\n\n" 
    
    if (pago.precio_aereo or 0) > 0 or (pago.cantidad_aereo or 0) > 0:
        content += f"  Precio Aereo: ¢{int(pago.precio_aereo or 0):,.0f}\n" \
                   f"  Cantidad Aereo: {int(pago.cantidad_aereo or 0)}\n" \
                   f"  Total General Aereo: ¢{int(total_general_aereo):,.0f}\n" \
                   f"  Total Individual Aereo: ¢{int(total_individual_aereo):,.0f}\n\n" 
    
    if (pago.precio_otros_transporte or 0) > 0 or (pago.cantidad_otros_transporte or 0) > 0 or pago.descripcion_otros_transporte:
        content += f"  Precio Otros Transporte: ¢{int(pago.precio_otros_transporte or 0):,.0f}\n" 
        if pago.descripcion_otros_transporte:
            content += f"  Descripción Otros Transporte: {pago.descripcion_otros_transporte}\n"
        content += f"  Cantidad Otros Transporte: {int(pago.cantidad_otros_transporte or 0)}\n" \
                   f"  Total General Otros: ¢{int(total_general_otros_transporte):,.0f}\n" \
                   f"  Total Individual Otros: ¢{int(total_individual_otros_transporte):,.0f}\n\n" 
    
    if total_general_transporte > 0:
        content += f"  Total General Transporte: ¢{int(total_general_transporte):,.0f}\n" \
                   f"  Total Individual Transporte: ¢{int(total_individual_transporte):,.0f}\n\n" 
    
    content += f"Otros Generales:\n" 
    if (pago.precio_guias or 0) > 0 or (pago.precio_guia_por_persona or 0) > 0 or (pago.cantidad_guias or 0) > 0:
        content += f"  Precio Guías (por grupo): ¢{int(pago.precio_guias or 0):,.0f}\n" \
                   f"  Precio Guías (por persona - referencia): ¢{int(pago.precio_guia_por_persona or 0):,.0f}\n" \
                   f"  Cantidad Guías: {int(pago.cantidad_guias or 0)}\n" \
                   f"  Total General Guías: ¢{int(total_general_guias):,.0f}\n" \
                   f"  Total Individual Guías: ¢{int(total_individual_guias):,.0f}\n\n" 
    
    content += f"Otros Personales:\n" 
    if (pago.precio_estadia or 0) > 0 or cantidad_dias_estadia_val > 0: # MODIFICADO
        content += f"  Precio Estadía (Base): ¢{int(pago.precio_estadia or 0):,.0f}\n" \
                   f"  Precio Estadía (Total por días): ¢{int(precio_estadia_calc_display):,.0f}\n" 
    if (pago.precio_impuestos or 0) > 0:
        content += f"  Precio Impuestos: ¢{int(pago.precio_impuestos or 0):,.0f}\n" 
    if (pago.precio_banos or 0) > 0:
        content += f"  Precio Baños: ¢{int(pago.precio_banos or 0):,.0f}\n" 
    if (pago.precio_servicios_sanitarios or 0) > 0:
        content += f"  Precio Servicios Sanitarios: ¢{int(pago.precio_servicios_sanitarios or 0):,.0f}\n" 
    if (pago.precio_acarreo or 0) > 0:
        content += f"  Precio Acarreo: ¢{int(pago.precio_acarreo or 0):,.0f}\n" 
    if (pago.precio_cafe or 0) > 0 or cantidad_dias_cafe_val > 0: # MODIFICADO
        content += f"  Precio Café (Base): ¢{int(pago.precio_cafe or 0):,.0f}\n" \
                   f"  Precio Café (Total por días): ¢{int(precio_cafe_calc_display):,.0f}\n" 
    if (pago.precio_cena or 0) > 0 or cantidad_dias_cena_val > 0: # MODIFICADO
        content += f"  Precio Cena (Base): ¢{int(pago.precio_cena or 0):,.0f}\n" \
                   f"  Precio Cena (Total por días): ¢{int(precio_cena_calc_display):,.0f}\n" 
    if (pago.precio_entrada or 0) > 0:
        content += f"  Precio Entrada: ¢{int(pago.precio_entrada or 0):,.0f}\n" 
    if (pago.precio_reconocimiento or 0) > 0:
        content += f"  Precio Reconocimiento: ¢{int(pago.precio_reconocimiento or 0):,.0f}\n" 
    if (pago.precio_permisos or 0) > 0:
        content += f"  Precio Permisos: ¢{int(pago.precio_permisos or 0):,.0f}\n" 
    if (pago.precio_pasaporte or 0) > 0:
        content += f"  Precio Pasaporte: ¢{int(pago.precio_pasaporte or 0):,.0f}\n" 
    
    if (pago.precio_otros1_personales or 0) > 0 or pago.descripcion_otros1_personales:
        content += f"  Precio Otros1: ¢{int(pago.precio_otros1_personales or 0):,.0f} ({pago.descripcion_otros1_personales or ''})\n"
    if (pago.precio_otros2_personales or 0) > 0 or pago.descripcion_otros2_personales:
        content += f"  Precio Otros2: ¢{int(pago.precio_otros2_personales or 0):,.0f} ({pago.descripcion_otros2_personales or ''})\n"
    if (pago.precio_otros3_personales or 0) > 0 or pago.descripcion_otros3_personales:
        content += f"  Precio Otros3: ¢{int(pago.precio_otros3_personales or 0):,.0f} ({pago.descripcion_otros3_personales or ''})\n"
    if (pago.precio_otros4_personales or 0) > 0 or pago.descripcion_otros4_personales:
        content += f"  Precio Otros4: ¢{int(pago.precio_otros4_personales or 0):,.0f} ({pago.descripcion_otros4_personales or ''})\n"

    if total_individual_personales_display > 0:
        content += f"  Total Individual Personales: ¢{int(total_individual_personales_display):,.0f}\n\n" 
    
    content += f"TOTALES FINALES:\n" \
              f"  TOTAL GENERAL: ¢{int(total_general_total_calc):,.0f}\n" \
              f"  TOTAL INDIVIDUAL: ¢{int(total_individual_total_calc):,.0f}\n" 
    if (pago.tipo_cambio or 0) > 0:
        content += f"  TOTAL INDIVIDUAL USD: ${total_individual_usd:,.2f}\n"
    
    content += f"  GANANCIA POR PERSONA: ¢{int(ganancia_pp):,.0f}\n" \
              f"  GANANCIA GENERAL: ¢{int(ganancia_gral):,.0f}\n\n"
    
    buffer = io.BytesIO()
    buffer.write(content.encode('utf-8'))
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='text/plain',
        as_attachment=True,
        download_name=f'detalle_pago_{pago.nombre_caminata.replace(" ", "_")}_{pago.id}.txt'
    )


def format_currency(value):
    return f"¢{value:,.0f}"

@pagos_bp.context_processor
def inject_format_currency():
    return dict(format_currency=format_currency)

